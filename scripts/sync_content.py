"""Sync Binance Square HTML exports into the Hugo site from an Excel control table.

The command is intentionally explicit: pass either ``--rows`` for a limited run or
``--all`` for a future full-table sync.  Dry-run mode performs parsing and remote
image validation in memory, but writes no Markdown, images, workbook cells, or log.
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import html as html_module
import io
import json
import os
import re
import shutil
import sys
import tempfile
import unicodedata
from copy import copy
from difflib import SequenceMatcher
from dataclasses import dataclass, field
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import unquote_to_bytes, urlparse
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from PIL import Image, ImageChops, ImageDraw, ImageFont, ImageOps
except ImportError as exc:  # pragma: no cover - dependency error is user-facing.
    raise SystemExit(
        "缺少运行依赖。请安装 openpyxl 和 Pillow：python -m pip install openpyxl Pillow"
    ) from exc


CATEGORY_MAP = {
    "板块研究": {
        "key": "sector",
        "section": "web3",
        "directory": "sector",
        "content_type": "sector-research",
    },
    "链上财报": {
        "key": "onchain-earnings",
        "section": "web3",
        "directory": "onchain-financials",
        "content_type": "onchain-earnings",
    },
    "项目分析": {
        "key": "project",
        "section": "web3",
        "directory": "projects",
        "content_type": "project-analysis",
    },
    "逻辑拆解": {
        "key": "thesis",
        "section": "web3",
        "directory": "thesis",
        "content_type": "thesis",
    },
    "财报分析": {
        "key": "earnings",
        "section": "stocks",
        "directory": "earnings",
        "content_type": "earnings-analysis",
    },
    "个股研究": {
        "key": "company",
        "section": "stocks",
        "directory": "company",
        "content_type": "company-research",
    },
    "产业报告": {
        "key": "industry",
        "section": "stocks",
        "directory": "industry",
        "content_type": "industry-report",
    },
}

PRIMARY_CATEGORY_LABELS = {"web3": "Web3", "stocks": "美股"}

ALLOWED_STATUSES = {"待导入", "已发布", "已归档", "处理失败"}
FEATURED_MAP = {"是": True, "否": False}
SEO_STATUSES = {"待生成", "已完成", "需重做", "处理失败"}
SEO_FIELDS = ["summary", "description", "key_points", "seo_keywords", "faq"]
REQUIRED_COLUMNS = ["ID", "HTML文件名", "Binance链接", "目标分类", "状态", "是否精选"]
SEO_STATUS_COLUMN = "SEO/GEO状态"
AUTHOR = "Eric SJ"
BANNED_SUMMARY_PREFIXES = ("本文从", "本文分析", "本文探讨", "本文围绕")
PROMPT_PHRASES = (
    "这是最核心的段落", "最核心的段落", "先看数据", "我觉得", "我认为",
    "下面来看", "下面我们看", "简单来说", "话不多说", "我们先看", "再来看",
    "接下来看看", "这一段", "上文提到", "前面提到",
)
BANNED_TEMPLATES = (*BANNED_SUMMARY_PREFIXES, "本文将", "本文主要", "本文介绍")

KEYWORD_VOCABULARY = [
    "Web3", "DeFi", "DePIN", "RWA", "AI", "GPU", "L1", "L2", "Rollup",
    "Bitcoin", "Ethereum", "Solana", "Base", "Arbitrum", "Spheron", "JBL",
    "Jabil", "GameStop", "游戏驿站", "GME", "eBay", "Hyperliquid", "Tether", "稳定币", "现金流", "商业模式", "收入",
    "利润", "财报", "估值", "人工智能", "基础设施", "算力", "去中心化",
    "项目分析", "板块研究", "财报分析", "个股研究", "产业报告", "链上财报",
]

NOISE_PHRASES = (
    "下载币安", "打开币安", "前往币安", "在币安广场", "查看更多推荐",
    "评论区", "相关推荐", "免责声明：", "风险提示：",
)


def clean_text(value: str | None) -> str:
    value = html_module.unescape(value or "")
    value = re.sub(r"[\u200b\ufeff]", "", value)
    return re.sub(r"\s+", " ", value).strip()


def yaml_string(value: str | None) -> str:
    return json.dumps(value or "", ensure_ascii=False)


def normalize_url(value: str | None) -> str:
    return clean_text(value).rstrip("/")


def normalize_title(value: str | None) -> str:
    """Return a strict comparison key without using fuzzy similarity."""
    value = unicodedata.normalize("NFKC", clean_text(value))
    value = re.sub(r"\.html?$", "", value, flags=re.I)
    value = re.sub(r"\s*\(\d+\)\s*$", "", value)
    value = re.sub(
        r"\s*[|｜]\s*Eric\s*SJ\s*发布于币安广场\s*\([^)]*\)\s*$",
        "",
        value,
        flags=re.I,
    )
    value = value.casefold()
    value = re.sub(r"\s+", "", value)
    value = value.replace("﹕", ":").replace("：", ":")
    return value.strip()


def normalize_status(value: str | None, matched_old: bool | None = None) -> str:
    """Return the Excel status unchanged after validation.

    ``matched_old`` is retained in the signature for compatibility with the
    normalization command, but Markdown existence must never change status.
    """
    raw = clean_text(value)
    if not raw:
        return "待导入"
    if raw not in ALLOWED_STATUSES:
        raise ValueError(f"无效状态：{raw or '（空）'}")
    return raw


def draft_for_status(status: str) -> bool:
    """Map the Excel-controlled lifecycle status to Hugo's draft flag."""
    if status not in ALLOWED_STATUSES:
        raise ValueError(f"无效状态：{status or '（空）'}")
    return status != "已发布"


def category_frontmatter(category_label: str) -> dict:
    category = CATEGORY_MAP[category_label]
    return {
        "category": PRIMARY_CATEGORY_LABELS[category["section"]],
        "subcategory": category_label,
        "content_type": category["content_type"],
        "category_key": category["key"],
        "category_label": category_label,
    }


def normalize_featured(value: str | None) -> tuple[str, bool]:
    normalized = "是" if clean_text(value) == "是" else "否"
    return normalized, normalized == "是"


class MetadataParser(HTMLParser):
    def __init__(self, key: str):
        super().__init__(convert_charrefs=True)
        self.key = key
        self.value = ""

    def handle_starttag(self, tag, attrs):
        values = dict(attrs)
        if tag == "meta" and (
            values.get("property") == self.key or values.get("name") == self.key
        ):
            self.value = values.get("content", "")


def meta(document: str, key: str) -> str:
    head_end = document.find("</head>")
    parser = MetadataParser(key)
    parser.feed(document[: head_end + 7 if head_end >= 0 else min(len(document), 150_000)])
    return html_module.unescape(parser.value)


def published_at(document: str) -> datetime | None:
    patterns = (
        r'"datePublished"\s*:\s*"([^"]+)',
        r'"publishedTime"\s*:\s*"([^"]+)',
        r'<time[^>]+datetime=["\']([^"\']+)',
    )
    for pattern in patterns:
        match = re.search(pattern, document, flags=re.I)
        if not match:
            continue
        raw = match.group(1).replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(raw).astimezone(ZoneInfo("Asia/Shanghai"))
        except ValueError:
            continue
    return None


def find_article_html(document: str) -> tuple[str, int]:
    start = document.find("class=richtext-container")
    if start < 0:
        start = document.find('class="richtext-container')
    if start < 0:
        raise ValueError("无法定位 Binance 正文容器 richtext-container")
    ends = [
        document.find('class="bottom-component-group', start),
        document.find("class=bottom-component-group", start),
        document.find('class="comment', start),
        document.find("class=comment", start),
    ]
    ends = [position for position in ends if position >= 0]
    end = min(ends) if ends else len(document)
    return document[start:end], start


class ArticleParser(HTMLParser):
    STRONG_OPEN = "\ue000"
    STRONG_CLOSE = "\ue001"
    BLOCKS = {"p", "h1", "h2", "h3", "h4", "li", "blockquote"}

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.blocks: list[tuple[str, str]] = []
        self.current: list | None = None
        self.strong_depth = 0
        self.list_stack: list[dict] = []
        self.link_stack: list[tuple[str, bool]] = []
        self.images: list[dict[str, str]] = []
        self.picture_depth = 0
        self.picture_sources: list[dict[str, str]] = []

    def finish(self):
        if not self.current:
            return
        kind, parts = self.current
        text = "".join(parts)
        pattern = re.compile(
            re.escape(self.STRONG_OPEN) + r"(.*?)" + re.escape(self.STRONG_CLOSE),
            flags=re.S,
        )
        while pattern.search(text):
            text = pattern.sub(
                lambda match: f"**{match.group(1).strip()}**"
                if match.group(1).strip()
                else "",
                text,
            )
        text = text.replace(self.STRONG_OPEN, "").replace(self.STRONG_CLOSE, "")
        text = re.sub(r"[ \t]+", " ", text).strip()
        if text and not any(phrase in text for phrase in NOISE_PHRASES):
            self.blocks.append((kind, text))
        self.current = None

    def handle_starttag(self, tag, attrs):
        values = dict(attrs)
        if tag == "picture":
            self.picture_depth += 1
            if self.picture_depth == 1:
                self.picture_sources = []
        elif tag == "source" and self.picture_depth:
            self.picture_sources.append({str(key): value or "" for key, value in values.items()})
        elif tag in {"ul", "ol"}:
            start = int(values.get("start", "1")) if str(values.get("start", "1")).isdigit() else 1
            self.list_stack.append({"tag": tag, "counter": start - 1})
        elif tag == "li":
            self.finish()
            if self.list_stack and self.list_stack[-1]["tag"] == "ol":
                self.list_stack[-1]["counter"] += 1
                self.current = [f"ol:{self.list_stack[-1]['counter']}", []]
            else:
                self.current = ["ul", []]
        elif tag in self.BLOCKS:
            self.finish()
            self.current = [tag, []]
        elif tag == "br" and self.current:
            self.current[1].append("\n")
        elif tag == "a" and self.current:
            href = values.get("href", "")
            parsed = urlparse(href)
            is_platform = "binance.com" in parsed.netloc.lower() or href.startswith("javascript:")
            self.link_stack.append((href, is_platform))
            if href and not is_platform:
                self.current[1].append("[")
        elif tag in {"strong", "b"} and self.current:
            self.current[1].append(self.STRONG_OPEN)
            self.strong_depth += 1
        elif tag == "img":
            self.finish()
            source_values = {str(key): value or "" for key, value in values.items()}
            source = source_values.get("src") or source_values.get("data-src") or ""
            if source or self.picture_sources:
                source_values["alt"] = clean_text(source_values.get("alt"))
                if self.picture_sources:
                    source_values["picture_sources"] = json.dumps(
                        self.picture_sources, ensure_ascii=False
                    )
                self.images.append(source_values)
                self.blocks.append(("img", str(len(self.images) - 1)))

    def handle_endtag(self, tag):
        if tag == "picture":
            self.picture_depth = max(0, self.picture_depth - 1)
            if not self.picture_depth:
                self.picture_sources = []
        elif tag == "li":
            self.finish()
        elif tag in {"ul", "ol"}:
            if self.list_stack:
                self.list_stack.pop()
        elif tag in self.BLOCKS:
            self.finish()
        elif tag == "a" and self.current and self.link_stack:
            href, is_platform = self.link_stack.pop()
            if href and not is_platform:
                self.current[1].append(f"]({href})")
        elif tag in {"strong", "b"} and self.current and self.strong_depth:
            self.current[1].append(self.STRONG_CLOSE)
            self.strong_depth -= 1

    def handle_data(self, data):
        if self.current:
            self.current[1].append(data)


def decode_data_image(source: str) -> bytes:
    header, encoded = source.split(",", 1)
    if ";base64" in header.lower():
        return base64.b64decode(encoded)
    if header.lower().startswith("data:image/svg+xml"):
        return unquote_to_bytes(encoded)
    raise ValueError("不支持的非 base64 data:image 资源")


def fetch_bytes(source: str) -> bytes:
    if source.startswith("data:image/"):
        return decode_data_image(source)
    request = Request(source, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=20) as response:
        return response.read()


def is_remote_source(source: str) -> bool:
    return source.startswith(("http://", "https://", "//"))


def css_variable_images(document: str) -> dict[str, str]:
    """Resolve SingleFile --sf-img-* CSS variables to their embedded data URI."""
    variables: dict[str, str] = {}
    # Avoid a DOTALL/back-reference scan across multi-megabyte SingleFile HTML.
    # Data URIs do not contain the surrounding quote, so two bounded patterns
    # are both faster and less prone to crossing unrelated CSS declarations.
    patterns = (
        re.compile(
            r'(--sf-img-[\w-]+)\s*:\s*url\(\s*"(data:image/[^\"]+)"\s*\)',
            flags=re.I,
        ),
        re.compile(
            r"(--sf-img-[\w-]+)\s*:\s*url\(\s*'(data:image/[^']+)'\s*\)",
            flags=re.I,
        ),
    )
    for pattern in patterns:
        for match in pattern.finditer(document):
            variables[match.group(1)] = html_module.unescape(
                match.group(2).replace("\\\n", "")
            )
    return variables


def embedded_data_uris(value: str | None) -> list[str]:
    value = html_module.unescape(value or "")
    results = re.findall(
        r"data:image/[a-z0-9.+-]+(?:;[a-z0-9=.+-]+)*;base64,[a-z0-9+/=\r\n]+",
        value,
        flags=re.I,
    )
    if value.strip().lower().startswith("data:image/svg+xml,"):
        results.append(value.strip().strip("\"'"))
    return list(dict.fromkeys(item.replace("\r", "").replace("\n", "") for item in results))


def css_image_sources(style: str | None, variables: dict[str, str]) -> list[str]:
    style = html_module.unescape(style or "")
    results: list[str] = []
    for name in re.findall(r"var\(\s*(--sf-img-[\w-]+)\s*\)", style, flags=re.I):
        if name in variables:
            results.append(variables[name])
    results.extend(embedded_data_uris(style))
    for match in re.finditer(r"url\(\s*([\"']?)([^)]+?)\1\s*\)", style, flags=re.I | re.S):
        source = clean_text(match.group(2)).strip("\"'")
        if source and not source.startswith("data:image/"):
            results.append(source)
    return list(dict.fromkeys(results))


def local_image_path(source: str, html_path: Path) -> Path | None:
    parsed = urlparse(source)
    if parsed.scheme or source.startswith(("//", "#", "blob:")):
        return None
    relative = html_module.unescape(parsed.path).replace("/", os.sep)
    if not relative:
        return None
    candidate = (html_path.parent / relative).resolve()
    try:
        candidate.relative_to(html_path.parent.resolve())
    except ValueError:
        return None
    return candidate if candidate.is_file() else None


def attribute_image_sources(
    attrs: dict[str, str], variables: dict[str, str], html_path: Path
) -> list[tuple[str, str]]:
    """Return element image sources in offline-first priority order."""
    candidates: list[tuple[int, str, str]] = []
    values = {str(key).lower(): value or "" for key, value in attrs.items()}
    src = values.get("src", "")

    def add(priority: int, source: str, kind: str):
        source = html_module.unescape(clean_text(source)).strip("\"'")
        if source:
            candidates.append((priority, source, kind))

    if src.startswith("data:image/") and "svg+xml" not in src[:80].lower():
        add(1, src, "data-uri")
    try:
        picture_sources = json.loads(values.get("picture_sources", "[]"))
    except json.JSONDecodeError:
        picture_sources = []
    for source_attrs in picture_sources:
        for key in ("srcset", "data-srcset", "src"):
            for source in embedded_data_uris(source_attrs.get(key)):
                add(2, source, "picture-source-data-uri")
        for source in css_image_sources(source_attrs.get("style"), variables):
            add(3, source, "picture-source-css")
    for key in ("srcset", "data-srcset"):
        for source in embedded_data_uris(values.get(key)):
            add(2, source, "srcset-data-uri")
    for source in css_image_sources(values.get("style"), variables):
        add(3, source, "inline-css")
    for key, value in values.items():
        if key in {"src", "srcset", "data-srcset", "style"}:
            continue
        for source in embedded_data_uris(value):
            add(4, source, "singlefile-resource-attribute")
    if src.startswith("data:image/"):
        add(5, src, "inline-svg" if "svg+xml" in src[:80].lower() else "data-uri")
    for key in ("data-src", "data-original-src", "data-lazy-src", "src"):
        value = values.get(key, "")
        if value and local_image_path(value, html_path):
            add(6, value, "local-file")
    for key in ("data-src", "data-original-src", "data-lazy-src", "src"):
        value = values.get(key, "")
        if is_remote_source(value):
            add(9, value, "remote")
    ordered = sorted(candidates, key=lambda item: item[0])
    seen: set[str] = set()
    result: list[tuple[str, str]] = []
    for _, source, kind in ordered:
        if source not in seen:
            seen.add(source)
            result.append((source, kind))
    return result


def read_image_source(
    source: str, kind: str, html_path: Path, allow_remote: bool = False
) -> bytes:
    if source.startswith("data:image/"):
        return decode_data_image(source)
    local_path = local_image_path(source, html_path)
    if local_path:
        return local_path.read_bytes()
    if is_remote_source(source):
        if not allow_remote:
            raise ValueError("离线模式未请求远程图片")
        return fetch_bytes(source)
    raise ValueError(f"无法解析图片来源：{kind}")


def image_info(data: bytes) -> tuple[int, int, str]:
    with Image.open(io.BytesIO(data)) as image:
        width, height = image.size
        format_name = (image.format or "PNG").lower().replace("jpeg", "jpg")
    return width, height, format_name


def img_attrs(tag: str) -> dict[str, str]:
    attrs = {}
    pattern = r"([:\w-]+)(?:\s*=\s*(\"[^\"]*\"|'[^']*'|[^\s>]+))?"
    for key, value in re.findall(pattern, tag):
        value = value or ""
        if len(value) >= 2 and value[0] in {"'", '"'} and value[-1] == value[0]:
            value = value[1:-1]
        attrs[key.lower()] = html_module.unescape(value)
    return attrs


def is_noise_image(attrs: dict[str, str], source: str) -> bool:
    marker = " ".join(
        [attrs.get("alt", ""), attrs.get("class", ""), attrs.get("aria-label", ""), source[:120]]
    ).lower()
    return source.startswith("data:image/svg") or any(
        word in marker for word in ("avatar", "logo", "icon", "emoji", "profile", "certified")
    )


@dataclass
class CoverCandidate:
    source: str
    source_kind: str
    region: str
    priority: int
    position: int


@dataclass
class CoverSelection:
    source: str
    source_label: str
    source_kind: str
    region: str
    candidate_count: int
    resource_count: int
    resource_types: dict[str, int]
    width: int
    height: int
    source_format: str
    is_true_top_visual: bool
    needs_normalization: bool
    output_width: int
    output_height: int
    output_format: str
    output_data: bytes
    used_remote: bool
    is_placeholder: bool
    cover_status: str
    read_errors: list[str]


def cover_candidates(
    document: str,
    article_start: int,
    title: str,
    html_path: Path,
    variables: dict[str, str],
) -> list[CoverCandidate]:
    """Return cover candidates without ever entering the article body container."""
    head_end = document.find("</head>")
    scan_start = head_end + 7 if head_end >= 0 else 0
    title_position = document.rfind(title, scan_start, article_start)
    candidates: list[CoverCandidate] = []

    # Binance long-form posts sometimes place a genuine hero as the very first
    # element inside richtext-container.  It is a cover only when no substantive
    # article text precedes it; the first chart/image after paragraphs is body
    # content and must never be promoted to cover.
    article_html, _ = find_article_html(document)
    article_end = article_start + len(article_html)
    container_content_start = document.find(">", article_start, article_end)
    if container_content_start >= 0:
        container_content_start += 1
        leading_img_start = re.search(
            r"<(?:img|source)\b",
            document[container_content_start:article_end],
            flags=re.I,
        )
        if leading_img_start:
            position = container_content_start + leading_img_start.start()
            prefix = document[container_content_start:position]
            prefix_text = clean_text(
                html_module.unescape(re.sub(r"<[^>]+>", " ", prefix))
            )
            if not prefix_text:
                tag_end = document.find(">", position, article_end)
                tag = document[position : tag_end + 1] if tag_end >= 0 else ""
                attrs = img_attrs(tag)
                for source, source_kind in attribute_image_sources(
                    attrs, variables, html_path
                ):
                    if not is_noise_image(attrs, source):
                        candidates.append(
                            CoverCandidate(
                                source,
                                source_kind,
                                "正文任何文字之前的顶部主视觉",
                                1,
                                position,
                            )
                        )

    for match in re.finditer(
        r"<(?:img|source)\b[^>]*>", document[scan_start:article_start], flags=re.I
    ):
        position = scan_start + match.start()
        attrs = img_attrs(match.group(0))
        sources = attribute_image_sources(attrs, variables, html_path)
        if not sources:
            continue
        marker = " ".join(
            (
                attrs.get("class", ""),
                attrs.get("id", ""),
                attrs.get("alt", ""),
                document[max(scan_start, position - 300):position],
            )
        ).lower()
        if any(
            word in marker
            for word in ("avatar", "profile", "certified", "author-info", "nickname")
        ):
            continue
        if title_position >= 0 and position > title_position:
            region = "文章标题之后、正文内容之前的顶部主视觉"
            priority = 1
        elif any(
            word in marker
            for word in (
                "hero",
                "cover",
                "banner",
                "article-header",
                "post-detail-article",
                "content-widget-buzz-post-detail",
                "css-mqxzqu",
            )
        ):
            region = "明确的文章头图区块"
            priority = 2
        else:
            continue
        for source, source_kind in sources:
            if not is_noise_image(attrs, source):
                candidates.append(
                    CoverCandidate(source, source_kind, region, priority, position)
                )

    for match in re.finditer(
        r"<[^>]+\bstyle\s*=\s*(?:\"[^\"]*(?:background|--sf-img-)[^\"]*\"|'[^']*(?:background|--sf-img-)[^']*'|[^\s>]*(?:background|--sf-img-)[^\s>]*)[^>]*>",
        document[scan_start:article_start],
        flags=re.I,
    ):
        position = scan_start + match.start()
        attrs = img_attrs(match.group(0))
        marker = " ".join(
            (attrs.get("class", ""), attrs.get("id", ""), document[max(scan_start, position - 300):position])
        ).lower()
        if any(
            word in marker
            for word in ("avatar", "profile", "certified", "author-info", "nickname")
        ):
            continue
        explicit_header = any(
            word in marker
            for word in ("hero", "cover", "banner", "article-header", "post-detail-article", "content-widget-buzz-post-detail")
        )
        if not explicit_header and not (title_position >= 0 and position > title_position):
            continue
        region = "明确的文章头图区块" if explicit_header else "文章标题之后、正文内容之前的顶部主视觉"
        priority = 2 if explicit_header else 1
        for source in css_image_sources(attrs.get("style"), variables):
            candidates.append(CoverCandidate(source, "inline-css", region, priority, position))

    # Binance may populate og:image with the first body chart.  Treat OG only as
    # a final fallback, and reject it when the same remote URL appears anywhere
    # in the actual article body.
    body_remote_sources = {
        html_module.unescape(source).split("?", 1)[0]
        for source in re.findall(
            r"https?://[^\s\"'<>),]+", document[article_start:article_end], flags=re.I
        )
    }

    og_image = clean_text(meta(document, "og:image"))
    og_key = html_module.unescape(og_image).split("?", 1)[0]
    if (
        og_image
        and bool(candidates)
        and og_key not in body_remote_sources
        and all(candidate.source != og_image for candidate in candidates)
    ):
        candidates.append(CoverCandidate(og_image, "remote", "og:image 备用图", 9, -1))

    candidates.sort(key=lambda item: (item.priority, -item.position))
    unique: list[CoverCandidate] = []
    seen: set[tuple[str, str]] = set()
    for candidate in candidates:
        key = (candidate.source, candidate.region)
        if key not in seen:
            seen.add(key)
            unique.append(candidate)
    return unique


def normalize_cover_contain(data: bytes) -> bytes:
    """Place a complete source image on a 1280x512 #faf5f0 canvas."""
    with Image.open(io.BytesIO(data)) as opened:
        source = ImageOps.exif_transpose(opened)
        source.thumbnail((1280, 512), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", (1280, 512), (250, 245, 240))
        left = (1280 - source.width) // 2
        top = (512 - source.height) // 2
        if source.mode in {"RGBA", "LA"} or (
            source.mode == "P" and "transparency" in source.info
        ):
            overlay = source.convert("RGBA")
            canvas.paste(overlay, (left, top), overlay.getchannel("A"))
        else:
            canvas.paste(source.convert("RGB"), (left, top))
        output = io.BytesIO()
        canvas.save(output, format="PNG", optimize=True)
        return output.getvalue()


def adapt_contain_cover_to_fill(data: bytes) -> tuple[bytes | None, tuple[int, int, int, int] | None]:
    """Convert our 5:2 contain canvas to a visually filled 5:2 cover.

    Only exact #faf5f0 pillar/letter boxes created by normalize_cover_contain
    are eligible. Native 5:2 artwork and deliberately padded designs remain
    untouched.
    """
    with Image.open(io.BytesIO(data)) as opened:
        image = ImageOps.exif_transpose(opened).convert("RGB")
        if image.size != (1280, 512):
            return None, None
        background = Image.new("RGB", image.size, (250, 245, 240))
        bbox = ImageChops.difference(image, background).getbbox()
        if not bbox:
            return None, None
        left, top, right, bottom = bbox
        vertical_span = top <= 2 and bottom >= 510
        horizontal_span = left <= 2 and right >= 1278
        pillarboxed = vertical_span and left >= 8 and (1280 - right) >= 8
        letterboxed = horizontal_span and top >= 8 and (512 - bottom) >= 8
        if not (pillarboxed or letterboxed):
            return None, bbox
        content = image.crop(bbox)
        fitted = ImageOps.fit(
            content,
            (1280, 512),
            method=Image.Resampling.LANCZOS,
            centering=(0.5, 0.5),
        )
        output = io.BytesIO()
        fitted.save(output, format="PNG", optimize=True)
        return output.getvalue(), bbox


def offline_resource_inventory(
    document: str, html_path: Path, variables: dict[str, str]
) -> dict[str, int]:
    inventory: dict[str, set[str]] = {
        "data-uri": set(),
        "inline-css": set(variables.values()),
        "srcset-data-uri": set(),
        "local-file": set(),
        "blob-reference": set(),
        "remote-reference": set(),
    }
    for source in embedded_data_uris(document):
        inventory["data-uri"].add(hashlib.sha1(source.encode("utf-8")).hexdigest())
    for match in re.finditer(r"\b(?:src|data-src|srcset)\s*=\s*([^\s>]+|\"[^\"]*\"|'[^']*')", document, flags=re.I):
        value = html_module.unescape(match.group(1).strip("\"'"))
        if "srcset" in match.group(0).lower():
            for source in embedded_data_uris(value):
                inventory["srcset-data-uri"].add(hashlib.sha1(source.encode("utf-8")).hexdigest())
        if value.startswith("blob:"):
            inventory["blob-reference"].add(value)
        elif is_remote_source(value):
            inventory["remote-reference"].add(value)
        elif local_image_path(value, html_path):
            inventory["local-file"].add(str(local_image_path(value, html_path)))
    return {key: len(value) for key, value in inventory.items() if value}


def placeholder_font(size: int):
    candidates = (
        Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "msyh.ttc",
        Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "simhei.ttf",
    )
    for path in candidates:
        if path.is_file():
            return ImageFont.truetype(str(path), size=size)
    return ImageFont.load_default()


def wrap_text_pixels(
    draw: ImageDraw.ImageDraw, text: str, font, maximum: int, maximum_lines: int = 4
) -> list[str]:
    lines: list[str] = []
    current = ""
    for char in text:
        candidate = current + char
        if current and draw.textbbox((0, 0), candidate, font=font)[2] > maximum:
            lines.append(current)
            current = char
        else:
            current = candidate
    if current:
        lines.append(current)
    if len(lines) > maximum_lines:
        lines = lines[:maximum_lines]
        last = lines[-1].rstrip("，。；：、 ")
        while last and draw.textbbox((0, 0), last + "…", font=font)[2] > maximum:
            last = last[:-1]
        lines[-1] = last + "…"
    return lines


def generate_placeholder_cover(title: str, category_label: str) -> bytes:
    canvas = Image.new("RGB", (1280, 512), (250, 245, 240))
    draw = ImageDraw.Draw(canvas)
    category_font = placeholder_font(24)
    title_font = placeholder_font(52)
    footer_font = placeholder_font(20)
    ink = (42, 40, 38)
    muted = (116, 106, 98)
    border = (226, 216, 207)

    category_box = draw.textbbox((0, 0), category_label, font=category_font)
    category_width = category_box[2] - category_box[0]
    draw.rounded_rectangle(
        (72, 62, 72 + category_width + 36, 108),
        radius=23,
        outline=border,
        width=2,
    )
    draw.text((90, 69), category_label, font=category_font, fill=muted)

    lines = wrap_text_pixels(draw, title, title_font, 1100, maximum_lines=2)
    line_height = 72
    title_top = 202 - (len(lines) - 1) * 36
    for index, line in enumerate(lines):
        draw.text((72, title_top + index * line_height), line, font=title_font, fill=ink)

    draw.line((72, 420, 1208, 420), fill=border, width=2)
    draw.text((72, 444), "Eric SJ Research", font=footer_font, fill=muted)
    output = io.BytesIO()
    canvas.save(output, format="PNG", optimize=True)
    return output.getvalue()


def select_cover(
    document: str,
    article_start: int,
    title: str,
    category_label: str,
    html_path: Path,
    allow_remote: bool = False,
    collect_inventory: bool = True,
) -> CoverSelection:
    variables = css_variable_images(document)
    inventory = (
        offline_resource_inventory(document, html_path, variables)
        if collect_inventory
        else {}
    )
    candidates = cover_candidates(document, article_start, title, html_path, variables)

    read_errors: list[str] = []
    for candidate in candidates:
        if candidate.source_kind == "remote" and not allow_remote:
            read_errors.append(f"{candidate.region}: 离线模式跳过远程请求")
            continue
        try:
            data = read_image_source(
                candidate.source, candidate.source_kind, html_path, allow_remote=allow_remote
            )
            width, height, source_format = image_info(data)
        except Exception as exc:
            read_errors.append(f"{candidate.region}: {exc}")
            continue
        ratio = width / height if height else 0
        if width < 600 or height < 180 or ratio < 1.2:
            read_errors.append(
                f"{candidate.region}: {width}x{height} 不是有效横版主视觉"
            )
            continue

        is_native_five_two = 2.48 <= ratio <= 2.52
        if is_native_five_two:
            output_data = data
            output_width, output_height = width, height
            output_format = source_format
        else:
            output_data = normalize_cover_contain(data)
            output_width, output_height = 1280, 512
            output_format = "png"
        return CoverSelection(
            source=candidate.source,
            source_label=(
                "嵌入式顶部图片（data:image）"
                if candidate.source.startswith("data:image/")
                else candidate.source
            ),
            source_kind=candidate.source_kind,
            region=candidate.region,
            candidate_count=len(candidates),
            resource_count=sum(inventory.values()),
            resource_types=inventory,
            width=width,
            height=height,
            source_format=source_format,
            is_true_top_visual=candidate.priority in {1, 2},
            needs_normalization=not is_native_five_two,
            output_width=output_width,
            output_height=output_height,
            output_format=output_format,
            output_data=output_data,
            used_remote=candidate.source_kind == "remote",
            is_placeholder=False,
            cover_status="正常",
            read_errors=read_errors,
        )

    placeholder = generate_placeholder_cover(title, category_label)
    return CoverSelection(
        source="",
        source_label="统一占位封面",
        source_kind="placeholder",
        region="HTML 内无可用顶部主视觉",
        candidate_count=len(candidates),
        resource_count=sum(inventory.values()),
        resource_types=inventory,
        width=0,
        height=0,
        source_format="",
        is_true_top_visual=False,
        needs_normalization=False,
        output_width=1280,
        output_height=512,
        output_format="png",
        output_data=placeholder,
        used_remote=False,
        is_placeholder=True,
        cover_status="缺失",
        read_errors=read_errors,
    )


def render_markdown(parser: ArticleParser, image_urls: list[str | None]) -> str:
    output: list[str] = []
    for kind, value in parser.blocks:
        if kind == "img":
            image_url = image_urls[int(value)]
            if image_url:
                output.append(f"![文章配图]({image_url})")
        elif kind in {"h1", "h2", "h3", "h4"}:
            level = 2 if kind in {"h1", "h2"} else 3
            output.append(f"{'#' * level} {value}")
        elif kind == "ul":
            output.append(f"- {value}")
        elif kind.startswith("ol:"):
            output.append(f"{kind.split(':', 1)[1]}. {value}")
        elif kind == "blockquote":
            output.append(f"> {value}")
        else:
            output.append(value)
    markdown = "\n\n".join(output).strip()
    markdown = re.sub(
        r"\[([^\]]+)\]\(https?://[^\s)]*binance\.com[^\s)]*\)", r"\1", markdown, flags=re.I
    )
    markdown = re.sub(r"https?://[^\s)]*binance\.com[^\s)]*", "", markdown, flags=re.I)
    markdown = re.sub(r"\*\*\s*\*\*", "", markdown)
    if markdown.count("**") % 2:
        raise ValueError("正文粗体标记不成对，检测到裸露的 **")
    for match in re.finditer(r"\*\*(.*?)\*\*", markdown, flags=re.S):
        if not match.group(1).strip():
            raise ValueError("正文包含空的 ** 粗体标记")
    return markdown


def render_structured_article_body(
    original_content: str,
) -> str:
    """Add only the body entry heading; templates render SEO/GEO modules."""
    return f"## 正文\n\n{original_content.strip()}".rstrip()


def plain_body_blocks(parser: ArticleParser) -> list[str]:
    blocks = []
    for kind, value in parser.blocks:
        if kind == "img" or kind.startswith("h"):
            continue
        value = re.sub(r"\*\*|\[([^\]]+)\]\([^)]*\)", lambda m: m.group(1) or "", value)
        value = clean_text(value)
        if len(value) >= 12:
            blocks.append(value)
    return blocks


def sentences(blocks: list[str]) -> list[str]:
    result: list[str] = []
    for block in blocks:
        for sentence in re.split(r"(?<=[。！？!?])\s*|\n+", block):
            sentence = clean_text(sentence).strip("-• ")
            if 18 <= len(sentence) <= 240 and sentence not in result:
                result.append(sentence)
    return result


def clip_sentence(value: str, maximum: int) -> str:
    value = clean_text(value)
    if len(value) <= maximum:
        return value
    clipped = value[:maximum]
    cut = max(clipped.rfind(mark) for mark in "，；、。！？")
    if cut >= max(20, maximum - 25):
        clipped = clipped[: cut + 1]
    else:
        clipped = clipped.rstrip("，；、： ") + "。"
    return clipped


def visible_length(value: str) -> int:
    return len(re.sub(r"\s+", "", clean_text(value)))


def clean_generated_text(value: str) -> str:
    value = clean_text(value).replace("**", "")
    value = value.replace("““", "“").replace("””", "”")
    value = re.sub(r"\.\s+(?=[\u4e00-\u9fff])", "。", value)
    value = re.sub(r"(?<=[%\u4e00-\u9fff])\.(?=[\u4e00-\u9fff])", "。", value)
    value = re.sub(r"(?<=[\u4e00-\u9fff])[,;:](?=[\u4e00-\u9fff])", lambda m: {",": "，", ";": "；", ":": "："}[m.group(0)], value)
    value = re.sub(r"[。]{2,}", "。", value)
    value = re.sub(r"[，]{2,}", "，", value)
    value = re.sub(r"[；]{2,}", "；", value)
    value = re.sub(r"[!！]{2,}", "！", value)
    value = re.sub(r"[?？]{2,}", "？", value)
    return value.strip("，；： ")


def ensure_terminal_punctuation(value: str) -> str:
    value = clean_generated_text(value).rstrip("，；、： ")
    if value and value[-1] not in "。！？!?":
        value += "。"
    return value


def similarity_text(value: str) -> str:
    value = clean_generated_text(value).casefold()
    return "".join(re.findall(r"[a-z0-9]+|[\u4e00-\u9fff]", value))


def text_similarity(left: str, right: str) -> float:
    left_key = similarity_text(left)
    right_key = similarity_text(right)
    if not left_key or not right_key:
        return 0.0
    return SequenceMatcher(None, left_key, right_key).ratio()


def has_abnormal_punctuation(value: str) -> bool:
    return bool(
        re.search(r"[，。；：！？,.!?;:]{2,}", value)
        or re.search(r"(?<=[%\u4e00-\u9fff])\.(?=[\u4e00-\u9fff])", value)
        or value.count("；") > 2
        or value.endswith(("，", "；", ":", "："))
    )


def contains_banned_language(value: str) -> bool:
    return any(phrase in value for phrase in (*BANNED_TEMPLATES, *PROMPT_PHRASES))


def sentence_score(value: str, title: str = "") -> int:
    signals = (
        "核心", "关键", "因为", "因此", "意味着", "决定", "收入", "利润", "增长",
        "需求", "市场", "竞争", "价值", "风险", "成本", "优势", "变化", "转向",
        "现金流", "商业模式", "护城河", "周期", "结构", "数据", "业务",
    )
    score = sum(3 for signal in signals if signal in value)
    score += 2 if 35 <= visible_length(value) <= 90 else 0
    title_terms = [term for term in re.findall(r"[A-Za-z][A-Za-z0-9.$-]{1,20}|[\u4e00-\u9fff]{2,6}", title) if len(term) >= 2]
    score += sum(1 for term in title_terms[:8] if term.casefold() in value.casefold())
    return score


def infer_subject(title: str, body_text: str) -> str:
    known = (
        ("JBL", "捷普（JBL）"),
        ("捷普", "捷普（JBL）"),
        ("老虎证券", "老虎证券"),
        ("Perp DEX", "永续合约DEX"),
        ("Coinbase", "Base"),
        ("Base", "Base"),
    )
    haystack = f"{title}\n{body_text}"
    for needle, subject in known:
        if needle.casefold() in haystack.casefold():
            return subject
    prefix = re.split(r"[：:丨|｜？?!！]", title, maxsplit=1)[0].strip("[]【】 ")
    if 2 <= visible_length(prefix) <= 18:
        return prefix
    entities = re.findall(r"[A-Za-z][A-Za-z0-9.$-]{1,18}", title)
    if entities:
        return entities[0]
    raise ValueError("无法从标题识别研究对象")


def synthesis_profile(title: str, category: str, blocks: list[str], attempt: int) -> dict | None:
    body_text = " ".join(blocks)
    haystack = f"{title}\n{body_text}".casefold()
    if "jbl" in haystack or "捷普" in haystack:
        if attempt == 1:
            return {
                "summary": "捷普正由传统电子制造代工转向AI数据中心基础设施供应商，收入与利润同步增长，但利润率、库存及客户集中度仍决定估值空间。",
                "description": "捷普的财报显示，AI数据中心、资本设备与网络通信需求正在改变其传统代工业务结构，收入和核心利润保持增长。文章同时检验利润率改善能否持续，并关注库存、大客户集中度及订单质量对后续估值的约束。",
                "key_points": [
                    "AI数据中心与网络通信需求正在提升捷普高增长业务在收入结构中的重要性。",
                    "核心利润增速快于收入增速，说明盈利质量较传统低毛利代工模式有所改善。",
                    "利润率能否达到管理层指引，是判断AI业务增长能否转化为股东回报的关键。",
                    "库存水平和大客户集中度仍可能放大订单波动与经营风险。",
                ],
            }
        return {
            "summary": "AI基础设施需求正在重塑捷普的业务结构，收入扩张已开始带动利润改善，但利润率兑现、库存压力与客户集中度仍是主要风险。",
            "description": "捷普正在摆脱单一电子代工厂的市场印象，AI数据中心、资本设备和网络通信业务成为新的增长来源。财报反映收入与核心利润同步改善，但盈利能否持续仍取决于利润率指引的兑现，以及库存和客户集中风险是否受控。",
            "key_points": [
                "AI基础设施相关业务已成为捷普收入增长的重要来源。",
                "利润增长快于收入增长，显示业务组合改善正在传导至盈利端。",
                "利润率指引的兑现程度将影响市场对捷普估值提升的判断。",
                "库存与客户集中度是观察经营韧性时不能忽视的风险变量。",
            ],
        }
    if "老虎证券" in haystack and ("罚款" in haystack or "转亏" in haystack):
        if attempt == 1:
            return {
                "summary": "老虎证券营收仍保持增长，但巨额监管罚款令季度利润由盈转亏，市场关注点已转向合规成本、业务修复与盈利恢复能力。",
                "description": "老虎证券本期收入延续增长，经纪、利息和增值服务仍提供业务支撑，但监管罚款显著侵蚀利润并造成季度亏损。文章据此评估成本扩张、合规压力和收入结构变化，判断公司能否恢复稳定盈利。",
                "key_points": [
                    "老虎证券收入保持增长，但监管罚款直接导致季度利润由盈转亏。",
                    "经纪、利息与增值服务构成主要收入来源，业务结构并非依赖单一板块。",
                    "成本增速高于收入增速，说明经营杠杆和费用控制仍需修复。",
                    "后续盈利恢复取决于合规风险收敛及核心业务增长能否覆盖新增成本。",
                ],
            }
        return {
            "summary": "监管罚款使老虎证券在收入增长的同时转为季度亏损，合规成本、费用控制与核心业务修复将决定盈利能否重新企稳。",
            "description": "老虎证券的收入端仍有增长，经纪、利息和增值服务提供了基本支撑，但一次性监管罚款和持续上升的成本令利润承压。文章重点判断合规问题能否收敛，以及业务增长能否覆盖费用并推动盈利恢复。",
            "key_points": [
                "收入增长未能抵消监管罚款影响，老虎证券当季由盈利转为亏损。",
                "经纪、利息和增值服务共同支撑收入，业务来源保持一定多样性。",
                "费用增长削弱了经营杠杆，成本控制将影响利润修复速度。",
                "监管风险是否收敛，是判断公司盈利稳定性的重要前提。",
            ],
        }
    if "perp dex" in haystack or ("永续合约" in haystack and "dex" in haystack):
        if attempt == 1:
            return {
                "summary": "永续合约DEX的竞争焦点正从链上交易能否成立，转向流动性留存、持仓深度与用户长期迁移，交易体验将决定下一阶段格局。",
                "description": "永续合约DEX已经越过早期功能验证阶段，竞争开始围绕未平仓量留存、流动性深度、交易体验和资金长期停留展开。文章同时比较中心化平台的信任成本，判断链上衍生品能否发展为独立交易所市场。",
                "key_points": [
                    "永续合约DEX的核心问题已从能否交易转向资金能否长期留在链上。",
                    "未平仓量留存和流动性深度比短期交易量更能反映平台竞争力。",
                    "中心化交易所的黑箱风险为链上衍生品平台提供了差异化机会。",
                    "交易体验、费用结构与持续流动性将共同决定行业格局。",
                ],
            }
        return {
            "summary": "链上永续合约市场已进入交易所竞争阶段，资金留存、流动性深度和使用体验比短期热度更重要，并将决定平台能否持续扩张。",
            "description": "链上永续合约不再只验证技术可行性，市场正在检验资金是否愿意长期保留仓位。文章聚焦未平仓量、流动性与交易体验，并结合中心化平台的信任成本，评估DEX争夺衍生品用户的长期空间。",
            "key_points": [
                "链上永续合约市场已经从产品验证进入资金留存竞争。",
                "未平仓量和流动性深度是判断平台真实使用度的重要指标。",
                "中心化平台的信任成本增强了透明交易基础设施的吸引力。",
                "产品体验和费用效率决定用户是否愿意长期迁移到链上。",
            ],
        }
    if "coinbase" in haystack and "base" in haystack:
        if attempt == 1:
            return {
                "summary": "Base的核心壁垒并非单纯依赖链上性能，而是Coinbase掌握的用户入口、分发能力与稳定币流动性共同形成增长飞轮。",
                "description": "Base在二层网络竞争中的优势，不只来自性能和费用，而取决于Coinbase能否把交易所用户入口、稳定币流动性与链上应用分发整合起来。文章据此评估用户转化、生态网络效应和持续收入能力。",
                "key_points": [
                    "Base的护城河主要来自Coinbase的用户入口和分发能力，而非单一技术指标。",
                    "稳定币流动性与链上应用入口可以共同强化Base的用户转化效率。",
                    "二层网络竞争正在从性能和费用比较转向争夺用户首次进入链上的入口。",
                    "Coinbase能否形成持续分发飞轮，将决定Base生态增长的质量。",
                ],
            }
        return {
            "summary": "Coinbase正在把交易所入口、稳定币流动性和应用分发能力注入Base，这种入口优势比单纯的性能竞争更可能形成长期壁垒。",
            "description": "Base的长期价值取决于Coinbase能否将既有用户、稳定币流动性和应用分发转化为链上增长。文章讨论的重点不是单纯比较二层网络性能，而是判断入口控制、生态转化与网络效应能否构成持续优势。",
            "key_points": [
                "Coinbase的用户入口为Base提供了其他二层网络难以复制的分发条件。",
                "稳定币流动性和应用入口有助于提升用户向链上场景的转化效率。",
                "Base的竞争逻辑更接近入口系统，而不是单纯追求更快或更便宜。",
                "生态增长能否形成网络效应，是判断其长期价值的核心变量。",
            ],
        }
    return None


def generic_synthesis(title: str, category: str, blocks: list[str], attempt: int) -> dict:
    body_text = " ".join(blocks)
    subject = infer_subject(title, body_text)
    themes = [
        term for term in (
            "收入增长", "利润改善", "现金流", "商业模式", "用户需求", "流动性",
            "竞争格局", "成本压力", "估值", "市场份额", "基础设施", "监管风险",
            "技术性能", "生态增长", "客户集中度", "库存风险",
        )
        if term.replace("增长", "") in body_text or term in body_text
    ]
    themes = list(dict.fromkeys(themes))
    if len(themes) < 3:
        raise ValueError("正文缺少足够的可验证主题，拒绝拼接原句生成SEO文案")
    if attempt == 1:
        summary = f"{subject}的{themes[0]}正在改变原有判断，{themes[1]}成为主要驱动因素，而{themes[2]}与{themes[-1]}仍决定长期表现和价值空间。"
        description = f"{subject}面临的核心问题，是{themes[0]}能否持续转化为{themes[1]}。研究范围同时覆盖{themes[2]}、{themes[-1]}与业务结构变化，用于判断现有优势是否具有持续性，以及主要风险会如何影响后续表现。"
    else:
        summary = f"围绕{subject}的重新定价正在转向基本面验证，{themes[0]}与{themes[1]}提供增长线索，但{themes[2]}和{themes[-1]}仍是判断持续性的关键。"
        description = f"{subject}的价值判断需要同时观察{themes[0]}、{themes[1]}和{themes[2]}。相关变化只有在业务结构与{themes[-1]}得到验证后，才能形成可持续优势；这些变量共同界定增长空间和潜在风险。"
    pool = [
        ensure_terminal_punctuation(clean_generated_text(item))
        for item in sorted(sentences(blocks), key=lambda item: -sentence_score(item, title))
        if not contains_banned_language(item) and not has_abnormal_punctuation(clean_generated_text(item))
    ]
    points = []
    for point in pool:
        if 20 <= visible_length(point) <= 120 and all(text_similarity(point, prior) < 0.8 for prior in points):
            points.append(point)
        if len(points) == 4:
            break
    if len(points) < 3:
        raise ValueError("无法提取3条完整且不重复的核心观点")
    return {"summary": summary, "description": description, "key_points": points}


def build_synthesis(title: str, category: str, blocks: list[str], attempt: int) -> dict:
    result = synthesis_profile(title, category, blocks, attempt)
    if result is None:
        result = generic_synthesis(title, category, blocks, attempt)
    result["summary"] = ensure_terminal_punctuation(result["summary"])
    result["description"] = ensure_terminal_punctuation(result["description"])
    result["key_points"] = [ensure_terminal_punctuation(point) for point in result["key_points"]]
    return result


def generate_key_points(blocks: list[str]) -> list[str]:
    pool = sentences(blocks)
    signals = ("核心", "因为", "意味着", "收入", "利润", "增长", "市场", "竞争", "价值", "风险", "成本", "优势", "变化")
    ranked = sorted(
        enumerate(pool),
        key=lambda item: (-sum(signal in item[1] for signal in signals), item[0]),
    )
    points: list[str] = []
    for _, candidate in ranked:
        candidate = ensure_terminal_punctuation(clip_sentence(candidate, 135))
        if candidate not in points:
            points.append(candidate)
        if len(points) >= 5:
            break
    points = points[:5]
    if len(points) < 3:
        raise ValueError("正文不足以提取3条真实核心观点")
    return points


def generate_keywords(title: str, blocks: list[str], category: str) -> list[str]:
    haystack = f"{title}\n{' '.join(blocks[:20])}"
    keywords: list[str] = []
    for word in KEYWORD_VOCABULARY:
        if word.lower() in haystack.lower() and word not in keywords:
            keywords.append(word)
    for token in re.findall(r"(?<![A-Za-z0-9])[A-Z][A-Z0-9.$-]{1,20}", haystack):
        token = token.strip(".$-")
        if token and token not in keywords:
            keywords.append(token)
    if category not in keywords:
        keywords.append(category)
    section_word = "Web3" if category in {"板块研究", "链上财报", "项目分析", "逻辑拆解"} else "美股"
    if section_word not in keywords:
        keywords.insert(0, section_word)
    fallback = ["行业研究", "商业模式", "价值分析", "市场研究", "投资研究"]
    for word in fallback:
        if len(keywords) >= 5:
            break
        if word not in keywords:
            keywords.append(word)
    return keywords[:8]


def generate_faq(title: str, category: str, summary: str, points: list[str], body: str) -> list[dict[str, str]]:
    if category in {"财报分析", "个股研究"}:
        return []
    suitable_category = category in {"板块研究", "链上财报", "逻辑拆解", "产业报告"}
    explanatory = bool(re.search(r"什么是|为什么|如何|解析|解读|商业模式|赛道|行业|原理", title + body[:1500]))
    if not (suitable_category or explanatory):
        return []
    clean_title = re.sub(r"[？?！!。]$", "", title)
    faqs = [{"question": f"{clean_title}的核心结论是什么？", "answer": summary}]
    if len(points) >= 2:
        faqs.append(
            {
                "question": "文章重点关注哪些关键变量？",
                "answer": clip_sentence(points[0] + points[1], 150),
            }
        )
    return faqs[:4]


def body_supports_point(point: str, body_text: str) -> bool:
    point_key = similarity_text(point)
    body_key = similarity_text(body_text)
    latin_terms = re.findall(r"[a-z][a-z0-9-]{1,20}", point_key)
    chinese = "".join(re.findall(r"[\u4e00-\u9fff]", point_key))
    bigrams = {chinese[index : index + 2] for index in range(max(0, len(chinese) - 1))}
    supported_bigrams = {term for term in bigrams if term in body_key}
    supported_latin = {term for term in latin_terms if term in body_key}
    return len(supported_bigrams) >= 3 or bool(supported_latin)


def validate_seo_geo(
    data: dict,
    blocks: list[str],
    title: str,
) -> dict:
    pool = sentences(blocks)
    first_sentence = pool[0] if pool else ""
    body_text = " ".join(blocks)
    summary = data.get("summary", "")
    description = data.get("description", "")
    points = data.get("key_points", [])
    keywords = data.get("seo_keywords", [])
    faq = data.get("faq", [])
    summary_first_similarity = text_similarity(summary, first_sentence)
    description_first_similarity = text_similarity(description, first_sentence)
    summary_title_similarity = text_similarity(summary, title)
    summary_sentence_similarity = max((text_similarity(summary, item) for item in pool), default=0.0)
    description_sentence_similarity = max((text_similarity(description, item) for item in pool), default=0.0)
    errors: list[str] = []

    if not 40 <= visible_length(summary) <= 70:
        errors.append(f"summary长度不合规：{visible_length(summary)}")
    if contains_banned_language(summary):
        errors.append("summary包含禁用模板或提示性表达")
    if has_abnormal_punctuation(summary):
        errors.append("summary存在异常标点")
    if summary_first_similarity >= 0.72:
        errors.append("summary与正文第一句高度相似")
    if summary_sentence_similarity >= 0.82:
        errors.append("summary疑似直接复制正文完整句子")
    if summary_title_similarity >= 0.72 or similarity_text(title) in similarity_text(summary):
        errors.append("summary明显重复标题")

    if not 80 <= visible_length(description) <= 160:
        errors.append(f"description长度不合规：{visible_length(description)}")
    if contains_banned_language(description) or "币安" in description:
        errors.append("description包含禁用模板、提示性表达或来源信息")
    if has_abnormal_punctuation(description):
        errors.append("description存在异常标点")
    if description_first_similarity >= 0.72:
        errors.append("description与正文第一句高度相似")
    if description_sentence_similarity >= 0.84:
        errors.append("description疑似直接复制正文句子")

    if not 3 <= len(points) <= 5:
        errors.append(f"key_points数量不合规：{len(points)}")
    for index, point in enumerate(points, start=1):
        if visible_length(point) < 18 or not point.endswith(("。", "！", "？")):
            errors.append(f"key_points第{index}条不是完整结论")
        if contains_banned_language(point):
            errors.append(f"key_points第{index}条包含提示性表达")
        if has_abnormal_punctuation(point):
            errors.append(f"key_points第{index}条存在异常标点")
        if not body_supports_point(point, body_text):
            errors.append(f"key_points第{index}条缺少正文依据")
    for index, point in enumerate(points):
        for prior in points[:index]:
            if text_similarity(point, prior) >= 0.8:
                errors.append(f"key_points第{index + 1}条与其他观点重复")
                break

    if not 5 <= len(keywords) <= 8 or len(set(keywords)) != len(keywords):
        errors.append("seo_keywords数量不合规或存在重复")
    if faq and not 2 <= len(faq) <= 4:
        errors.append("faq必须为2至4组或空数组")
    if any(contains_banned_language(f"{item.get('question', '')}{item.get('answer', '')}") for item in faq):
        errors.append("faq包含提示性表达")

    return {
        "passed": not errors,
        "errors": errors,
        "summary_length": visible_length(summary),
        "description_length": visible_length(description),
        "summary_first_sentence_similarity": round(summary_first_similarity, 4),
        "description_first_sentence_similarity": round(description_first_similarity, 4),
        "summary_title_similarity": round(summary_title_similarity, 4),
        "summary_max_body_sentence_similarity": round(summary_sentence_similarity, 4),
        "description_max_body_sentence_similarity": round(description_sentence_similarity, 4),
        "first_sentence_highly_similar": summary_first_similarity >= 0.72 or description_first_similarity >= 0.72,
    }


def generate_seo_geo(blocks: list[str], title: str, category: str, body: str) -> dict:
    attempts: list[dict] = []
    for attempt in (1, 2):
        synthesis = build_synthesis(title, category, blocks, attempt)
        seo_keywords = generate_keywords(title, blocks, category)
        faq = generate_faq(
            title,
            category,
            synthesis["summary"],
            synthesis["key_points"],
            body,
        )
        result = {
            "summary": synthesis["summary"],
            "description": synthesis["description"],
            "key_points": synthesis["key_points"],
            "seo_keywords": seo_keywords,
            "faq": faq,
        }
        quality = validate_seo_geo(result, blocks, title)
        attempts.append({"attempt": attempt, **quality})
        if quality["passed"]:
            result["_quality"] = quality
            result["_attempts"] = attempts
            result["_final_status"] = "已完成"
            return result
    raise ValueError(
        "两次SEO/GEO生成均未通过质量校验："
        + json.dumps(attempts, ensure_ascii=False)
    )


def failed_seo_geo(reason: str) -> dict:
    """Return deliberately empty SEO/GEO fields without blocking article import."""
    return {
        "summary": "",
        "description": "",
        "key_points": [],
        "seo_keywords": [],
        "faq": [],
        "_quality": {
            "passed": False,
            "errors": [reason],
            "first_sentence_highly_similar": False,
        },
        "_attempts": [],
        "_final_status": "处理失败",
        "_error": reason,
    }


WINDOWS_RESERVED_NAMES = {
    "con",
    "prn",
    "aux",
    "nul",
    *(f"com{number}" for number in range(1, 10)),
    *(f"lpt{number}" for number in range(1, 10)),
}
MAX_SLUG_LENGTH = 80
MAX_WINDOWS_PATH_LENGTH = 240


def sanitize_slug(value: str) -> str:
    """Return a lowercase ASCII slug that is safe as a Windows path component."""
    normalized = unicodedata.normalize("NFKC", value).lower()
    slug = "-".join(re.findall(r"[a-z0-9]+", normalized))
    slug = re.sub(r"-+", "-", slug).strip(" .-")[:MAX_SLUG_LENGTH].strip(" .-")
    if slug in WINDOWS_RESERVED_NAMES:
        slug = f"article-{slug}"
    return slug


def slug_validation_errors(slug: str) -> list[str]:
    errors: list[str] = []
    if not slug:
        errors.append("slug is empty")
    if slug != slug.lower():
        errors.append("slug is not lowercase")
    if not re.fullmatch(r"[a-z0-9]+(?:-[a-z0-9]+)*", slug or ""):
        errors.append("slug contains characters outside lowercase ASCII, digits, and hyphens")
    if slug.endswith((".", " ", "-")) or slug.startswith((".", " ", "-")):
        errors.append("slug starts or ends with a dot, space, or hyphen")
    if re.search(r"[<>:\"/\\|?*\x00-\x1f]", slug):
        errors.append("slug contains a Windows-invalid character")
    if slug.lower() in WINDOWS_RESERVED_NAMES:
        errors.append("slug is a reserved Windows device name")
    if len(slug) > MAX_SLUG_LENGTH:
        errors.append(f"slug exceeds {MAX_SLUG_LENGTH} characters")
    return errors


def ensure_safe_article_slug(
    candidate: str,
    source_url: str,
    site_root: Path,
    content_dir: Path,
    section: str,
    date_prefix: str,
) -> str:
    """Validate the Markdown filename and eventual Hugo public path before writing."""
    slug = sanitize_slug(candidate)
    post_id = re.search(r"/post/(\d+)", source_url)
    fallback = sanitize_slug(
        f"{date_prefix}-binance-{post_id.group(1) if post_id else 'article'}"
    )

    def planned_paths(value: str) -> tuple[Path, Path]:
        markdown_path = content_dir / f"{value}.md"
        public_path = site_root / "public" / "zh-cn" / section / value / "index.html"
        return markdown_path, public_path

    markdown_path, public_path = planned_paths(slug)
    errors = slug_validation_errors(slug)
    if len(str(markdown_path.resolve())) > MAX_WINDOWS_PATH_LENGTH:
        errors.append("Markdown path is too long")
    if len(str(public_path.resolve())) > MAX_WINDOWS_PATH_LENGTH:
        errors.append("public path is too long")

    if errors:
        slug = fallback
        markdown_path, public_path = planned_paths(slug)
        errors = slug_validation_errors(slug)
        if len(str(markdown_path.resolve())) > MAX_WINDOWS_PATH_LENGTH:
            errors.append("fallback Markdown path is too long")
        if len(str(public_path.resolve())) > MAX_WINDOWS_PATH_LENGTH:
            errors.append("fallback public path is too long")
    if errors:
        raise ValueError(f"Unable to generate a Windows-safe slug: {'; '.join(errors)}")
    return slug


def slugify(title: str, source_url: str) -> str:
    normalized = unicodedata.normalize("NFKC", title).lower()
    tokens = re.findall(r"[a-z0-9]+", normalized)
    aliases = []
    named_aliases = (
        ("游戏驿站", ["gamestop", "earnings"]),
        ("老虎证券", ["tiger-brokers", "earnings"]),
        ("美光", ["micron", "mu"]),
        ("英伟达", ["nvidia", "nvda"]),
    )
    for needle, values in named_aliases:
        if needle in title:
            aliases += values
    if "bio" in tokens and "上线之前" in title:
        aliases += ["bio", "before", "launch"]
    if "人工智能" in title:
        aliases += ["ai", "infrastructure"]
    if "算力" in title:
        aliases += ["compute"]
    if "去中心化" in title:
        aliases += ["decentralized"]
    words: list[str] = []
    for word in [*aliases, *tokens]:
        if word not in words and word not in {"eric", "sj"}:
            words.append(word)
    post_id = re.search(r"/post/(\d+)", source_url)
    if not words:
        words = ["binance", post_id.group(1) if post_id else "post"]
    return sanitize_slug("-".join(words[:7]))


def split_frontmatter(text: str) -> tuple[str, str, str]:
    newline = "\r\n" if "\r\n" in text[:500] else "\n"
    if not text.startswith(f"---{newline}"):
        raise ValueError("Markdown 缺少 YAML Front Matter")
    marker = f"{newline}---{newline}"
    end = text.find(marker, 4)
    if end < 0:
        raise ValueError("Markdown Front Matter 未闭合")
    return text[len(f"---{newline}") : end], text[end + len(marker) :], newline


def scalar_from_frontmatter(frontmatter: str, key: str) -> str | bool | None:
    match = re.search(rf"(?m)^{re.escape(key)}:[ \t]*([^\r\n]*?)[ \t]*\r?$", frontmatter)
    if not match:
        return None
    raw = match.group(1).strip()
    if raw in {"true", "false"}:
        return raw == "true"
    try:
        value = json.loads(raw)
        return value if isinstance(value, (str, bool)) else raw
    except json.JSONDecodeError:
        return raw.strip("'\"")


def frontmatter_field_span(frontmatter: str, key: str) -> tuple[int, int] | None:
    lines = frontmatter.splitlines(keepends=True)
    offset = 0
    start = None
    for index, line in enumerate(lines):
        if re.match(rf"^{re.escape(key)}\s*:", line):
            start = offset
            end = start + len(line)
            for following in lines[index + 1 :]:
                if re.match(r"^[A-Za-z_][A-Za-z0-9_/-]*\s*:", following):
                    break
                end += len(following)
            return start, end
        offset += len(line)
    return None


def frontmatter_field_raw(frontmatter: str, key: str) -> str | None:
    span = frontmatter_field_span(frontmatter, key)
    return frontmatter[slice(*span)].rstrip("\r\n") if span else None


def list_item_count(raw: str | None, child_key: str | None = None) -> int:
    if not raw:
        return 0
    if child_key:
        return len(re.findall(rf"(?m)^\s+-\s+{re.escape(child_key)}\s*:", raw))
    return len(re.findall(r"(?m)^\s+-\s+", raw))


def seo_geo_complete(frontmatter: str) -> bool:
    summary = clean_text(str(scalar_from_frontmatter(frontmatter, "summary") or ""))
    description = clean_text(str(scalar_from_frontmatter(frontmatter, "description") or ""))
    key_points = frontmatter_field_raw(frontmatter, "key_points")
    keywords = frontmatter_field_raw(frontmatter, "seo_keywords")
    faq = frontmatter_field_raw(frontmatter, "faq")
    faq_valid = bool(faq and (re.search(r"^faq:\s*\[\s*\]\s*$", faq) or list_item_count(faq, "question") >= 1))
    return bool(
        summary
        and description
        and list_item_count(key_points) >= 1
        and list_item_count(keywords) >= 1
        and faq_valid
    )


def seo_field_complete(frontmatter: str, key: str) -> bool:
    if key in {"summary", "description"}:
        return bool(clean_text(str(scalar_from_frontmatter(frontmatter, key) or "")))
    raw = frontmatter_field_raw(frontmatter, key)
    if key == "key_points":
        return list_item_count(raw) >= 1
    if key == "seo_keywords":
        return list_item_count(raw) >= 1
    if key == "faq":
        return bool(raw and (re.search(r"^faq:\s*\[\s*\]\s*$", raw) or list_item_count(raw, "question") >= 1))
    return False


def normalize_seo_status(value: str | None, article: "ExistingArticle | None") -> str:
    raw = clean_text(value)
    if raw == "需重做":
        return "需重做"
    if article and article.seo_complete:
        return "已完成"
    if raw == "处理失败":
        return "处理失败"
    return "待生成"


def markdown_body_blocks(body: str) -> list[str]:
    body = re.sub(r"!\[[^\]]*\]\([^)]*\)", "", body)
    body = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", body)
    body = body.replace("**", "").replace("__", "")
    blocks: list[str] = []
    for chunk in re.split(r"\n\s*\n", body):
        lines = []
        for line in chunk.splitlines():
            line = re.sub(r"^\s{0,3}(?:#{1,6}|>|[-*+] |\d+[.)] )\s*", "", line)
            line = clean_text(line)
            if line and not line.startswith("/images/"):
                lines.append(line)
        value = clean_text(" ".join(lines))
        if len(value) >= 12:
            blocks.append(value)
    return blocks


def render_frontmatter_field(key: str, value, newline: str) -> str:
    if isinstance(value, bool):
        return f"{key}: {str(value).lower()}"
    if isinstance(value, list):
        if not value:
            return f"{key}: []"
        lines = [f"{key}:"]
        for item in value:
            if isinstance(item, dict):
                lines.append(f"  - question: {yaml_string(str(item['question']))}")
                lines.append(f"    answer: {yaml_string(str(item['answer']))}")
            else:
                lines.append(f"  - {yaml_string(str(item))}")
        return newline.join(lines)
    return f"{key}: {yaml_string(str(value))}"


def field_report_value(frontmatter: str, key: str):
    raw = frontmatter_field_raw(frontmatter, key)
    if key in {"key_points", "seo_keywords", "faq"} and raw:
        if key == "faq":
            if re.search(r"^faq:\s*\[\s*\]\s*$", raw):
                return []
            return f"{list_item_count(raw, 'question')}组"
        return f"{list_item_count(raw)}项"
    scalar = scalar_from_frontmatter(frontmatter, key)
    if scalar is not None:
        return scalar
    if not raw:
        return None
    if key == "faq":
        if re.search(r"^faq:\s*\[\s*\]\s*$", raw):
            return []
        return f"{list_item_count(raw, 'question')}组"
    return f"{list_item_count(raw)}项"


def update_frontmatter_only(path: Path, values: dict, dry_run: bool) -> dict[str, tuple]:
    text = path.read_bytes().decode("utf-8")
    frontmatter, body, newline = split_frontmatter(text)
    changes = {}
    updated = frontmatter
    for key, value in values.items():
        old = field_report_value(frontmatter, key)
        changes[key] = (old, value)
        rendered = render_frontmatter_field(key, value, newline)
        span = frontmatter_field_span(updated, key)
        if span:
            prefix = updated[: span[0]]
            suffix = updated[span[1] :]
            if prefix and not prefix.endswith(("\n", "\r")):
                prefix += newline
            if suffix and not suffix.startswith(("\n", "\r")):
                rendered += newline
            updated = prefix + rendered + suffix
        else:
            updated = updated.rstrip("\r\n") + newline + rendered
    if not dry_run:
        result = f"---{newline}{updated}{newline}---{newline}{body}"
        path.write_bytes(result.encode("utf-8"))
    return {**changes, "_body_sha256": (hashlib.sha256(body.encode("utf-8")).hexdigest(), hashlib.sha256(body.encode("utf-8")).hexdigest())}


@dataclass
class ExistingArticle:
    path: Path
    content_id: str
    title: str
    source_url: str
    frontmatter: str
    body: str
    seo_complete: bool
    virtual: bool = False


def content_index(
    content_root: Path,
) -> tuple[
    dict[str, list[ExistingArticle]],
    dict[str, list[ExistingArticle]],
    dict[str, list[ExistingArticle]],
]:
    by_id: dict[str, list[ExistingArticle]] = {}
    by_url: dict[str, list[ExistingArticle]] = {}
    by_title: dict[str, list[ExistingArticle]] = {}
    for path in content_root.glob("**/*.md"):
        if path.name.startswith("_") or path.stem == "template":
            continue
        try:
            text = path.read_bytes().decode("utf-8")
            frontmatter, body, _newline = split_frontmatter(text)
        except (UnicodeDecodeError, ValueError):
            continue
        content_id = clean_text(str(scalar_from_frontmatter(frontmatter, "content_id") or ""))
        title = clean_text(str(scalar_from_frontmatter(frontmatter, "title") or ""))
        source_url = normalize_url(str(scalar_from_frontmatter(frontmatter, "source_url") or ""))
        article = ExistingArticle(
            path,
            content_id,
            title,
            source_url,
            frontmatter,
            body,
            seo_geo_complete(frontmatter),
        )
        if content_id:
            by_id.setdefault(content_id, []).append(article)
        if source_url:
            by_url.setdefault(source_url, []).append(article)
        if title:
            by_title.setdefault(normalize_title(title), []).append(article)
    return by_id, by_url, by_title


def register_article_in_indexes(
    article: ExistingArticle,
    by_id: dict[str, list[ExistingArticle]],
    by_url: dict[str, list[ExistingArticle]],
    by_title: dict[str, list[ExistingArticle]],
) -> None:
    """Update live indexes so dry-run and real runs make identical match decisions."""
    if article.content_id:
        by_id.setdefault(article.content_id, []).append(article)
    if article.source_url:
        by_url.setdefault(article.source_url, []).append(article)
    if article.title:
        by_title.setdefault(normalize_title(article.title), []).append(article)


def markdown_body_blocks(body: str) -> list[str]:
    """Read useful prose from Markdown without changing or re-parsing the source HTML."""
    blocks: list[str] = []
    for raw in body.splitlines():
        line = raw.strip()
        if not line or line.startswith(("![", "```", "---")):
            continue
        line = re.sub(r"^#{1,6}\s+", "", line)
        line = re.sub(r"^>\s*", "", line)
        line = re.sub(r"^[-*+]\s+", "", line)
        line = re.sub(r"^\d+[.)]\s+", "", line)
        line = re.sub(r"\*\*|__|`", "", line)
        line = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", line)
        line = clean_text(line)
        if visible_length(line) >= 12:
            blocks.append(line)
    return blocks


def markdown_first_sentence(body: str) -> str:
    blocks = markdown_body_blocks(body)
    if not blocks:
        return ""
    parts = re.split(r"(?<=[。！？!?])\s*", blocks[0], maxsplit=1)
    return clean_text(parts[0])


def summary_copy_quality(summary: str, description: str, title: str, body: str) -> dict:
    blocks = markdown_body_blocks(body)
    body_text = " ".join(blocks)
    pool = sentences(blocks)
    first_sentence = markdown_first_sentence(body)
    summary = clean_generated_text(summary)
    description = clean_generated_text(description)
    errors: list[str] = []
    summary_first_similarity = text_similarity(summary, first_sentence)
    description_first_similarity = text_similarity(description, first_sentence)
    summary_body_similarity = max((text_similarity(summary, item) for item in pool), default=0.0)
    description_body_similarity = max(
        (text_similarity(description, item) for item in pool), default=0.0
    )
    if not 40 <= visible_length(summary) <= 70:
        errors.append(f"summary长度不合规：{visible_length(summary)}")
    if contains_banned_language(summary):
        errors.append("summary包含禁用模板或提示性表达")
    if has_abnormal_punctuation(summary):
        errors.append("summary存在异常标点")
    if summary_first_similarity >= 0.72:
        errors.append("summary与正文第一句高度相似")
    if summary_body_similarity >= 0.82:
        errors.append("summary疑似复制正文完整句子")
    if not body_supports_point(summary, body_text):
        errors.append("summary缺少正文依据")
    if text_similarity(summary, title) >= 0.72 or similarity_text(title) in similarity_text(summary):
        errors.append("summary明显重复标题")
    if not 80 <= visible_length(description) <= 160:
        errors.append(f"description长度不合规：{visible_length(description)}")
    if contains_banned_language(description) or "Binance" in description or "币安广场" in description:
        errors.append("description包含禁用模板、提示语或来源信息")
    if has_abnormal_punctuation(description):
        errors.append("description存在异常标点")
    if description_first_similarity >= 0.72:
        errors.append("description与正文第一句高度相似")
    if description_body_similarity >= 0.84:
        errors.append("description疑似复制正文完整句子")
    if not body_supports_point(description, body_text):
        errors.append("description缺少正文依据")
    return {
        "passed": not errors,
        "errors": errors,
        "first_sentence": first_sentence,
        "summary_length": visible_length(summary),
        "description_length": visible_length(description),
        "summary_first_sentence_similarity": round(summary_first_similarity, 4),
        "description_first_sentence_similarity": round(description_first_similarity, 4),
        "summary_max_body_sentence_similarity": round(summary_body_similarity, 4),
        "description_max_body_sentence_similarity": round(description_body_similarity, 4),
        "first_sentence_highly_similar": summary_first_similarity >= 0.72,
    }


def frontmatter_list_values(frontmatter: str, key: str) -> list[str]:
    """Read a flat YAML list emitted by this script without loading the whole document."""
    raw = frontmatter_field_raw(frontmatter, key)
    if not raw:
        return []
    values: list[str] = []
    for match in re.finditer(r"(?m)^\s+-\s+(.+?)\s*$", raw):
        value = match.group(1).strip()
        if value.startswith("question:"):
            continue
        if value.startswith('"'):
            try:
                value = str(json.loads(value))
            except json.JSONDecodeError:
                value = value.strip('"')
        elif value.startswith("'") and value.endswith("'"):
            value = value[1:-1].replace("''", "'")
        value = clean_generated_text(value)
        if value:
            values.append(value)
    return values


def core_points_quality(points: list[str], title: str, body: str) -> dict:
    blocks = markdown_body_blocks(body)
    body_text = " ".join(blocks)
    errors: list[str] = []
    normalized = [ensure_terminal_punctuation(point) for point in points]
    process_phrases = (
        "本篇", "本文", "这篇文章", "这篇帖子", "通篇", "我个人", "我知道",
        "我认为", "我觉得", "大家", "朋友", "下面", "上图", "下图", "原推",
        "介绍完", "聊聊", "先说", "先看", "再看", "接下来", "欢迎讨论",
    )
    if not 3 <= len(normalized) <= 5:
        errors.append(f"核心观点数量不合规：{len(normalized)}")
    for index, point in enumerate(normalized, start=1):
        if not 18 <= visible_length(point) <= 140:
            errors.append(f"核心观点第{index}条长度不合规：{visible_length(point)}")
        if contains_banned_language(point):
            errors.append(f"核心观点第{index}条包含写作提示语")
        if any(phrase in point for phrase in process_phrases):
            errors.append(f"核心观点第{index}条包含写作过程或口语化表达")
        if has_abnormal_punctuation(point):
            errors.append(f"核心观点第{index}条存在异常标点")
        if "？" in point or "?" in point:
            errors.append(f"核心观点第{index}条是问题而非结论")
        if re.match(r"^\s*(?:\d+[.、）)]|[（(]\d+[）)])", point):
            errors.append(f"核心观点第{index}条包含正文编号残留")
        if point.rstrip("。！").endswith(("假如", "比如", "因为", "但是", "不过", "两个", "如下")):
            errors.append(f"核心观点第{index}条语义不完整")
        if text_similarity(point, title) >= 0.82:
            errors.append(f"核心观点第{index}条重复标题")
        if not body_supports_point(point, body_text):
            errors.append(f"核心观点第{index}条缺少正文依据")
    for index, point in enumerate(normalized):
        if any(text_similarity(point, prior) >= 0.8 for prior in normalized[:index]):
            errors.append(f"核心观点第{index + 1}条与其他观点重复")
    return {
        "passed": not errors,
        "errors": errors,
        "count": len(normalized),
        "points": normalized,
    }


def core_points_profile(title: str) -> list[str] | None:
    profiles = (
        (
            ("英伟达", "NVDA"),
            [
                "数据中心仍是英伟达最主要的增长引擎，Blackwell放量继续支撑高端GPU收入。",
                "整体营收仍保持增长，但增速已较此前高位明显放缓。",
                "毛利率回落主要与Blackwell新品爬坡和供应链成本有关，而非需求突然消失。",
                "推理模型提升计算需求，DeepSeek带来的效率变化尚不足以证明GPU需求进入平台期。",
            ],
        ),
        (
            ("BIO上线", "BIO Protocol"),
            [
                "BIO Protocol通过BioDAO连接患者、科学家和资金，为早期科研项目提供融资与协作网络。",
                "知识产权代币化使科研成果能够在链上确权、治理并引入更广泛的资金参与。",
                "DeFi与Web3工具降低了科研社区融资和资源协调的门槛。",
                "BIO的长期价值取决于科研成果转化、社区治理效率与真实项目供给能否持续。",
            ],
        ),
        (
            ("以太 $ETH", "以太坊价值"),
            [
                "高额交易费用既反映以太坊区块空间存在真实需求，也提高了普通用户的使用门槛。",
                "Layer2承担扩容任务，但主网仍需维持安全结算和数据可用性的核心地位。",
                "ETH的价值来源包括主网安全、生态网络效应、质押需求和区块空间收费。",
                "长期价值取决于扩容红利能否回流主网，并形成足以支持安全预算的价值捕获。",
            ],
        ),
        (
            ("PMF",),
            [
                "Web3产品不能只凭空投、补贴或短期活跃判断是否已经找到PMF。",
                "真实需求、持续使用、用户留存与付费意愿需要同时得到验证。",
                "降低跨链和交互门槛能够改善产品体验，但不能替代真实市场需求。",
                "依赖外部激励维持的增长难以持续，能够形成自然使用循环的产品才更接近PMF。",
            ],
        ),
        (
            ("双代币", "积分+代币"),
            [
                "双代币模型将治理价值与高频使用价值分开，能够降低二级市场波动对生态使用场景的干扰。",
                "积分与代币模式可根据需求调整兑换比例，从而延长激励周期并平衡供需。",
                "双向兑换可能通过套利和通缩机制促进流通，但也显著提高机制设计复杂度。",
                "多种资产并存会增加用户理解成本，并可能进一步分散市场注意力和流动性。",
            ],
        ),
        (
            ("DePIN丨Roam",),
            [
                "Roam通过WiFi节点、eSIM和代币激励连接无线网络供给方与使用者。",
                "设备数量领先说明网络扩张较快，但规模本身不能直接证明商业模式已经成立。",
                "免费连接和代币补贴有助于冷启动，也使网络增长对激励投入形成依赖。",
                "长期价值取决于补贴退坡后能否保留真实用户，并形成可持续的资源供给与收入。",
            ],
        ),
    )
    for needles, points in profiles:
        if any(needle.casefold() in title.casefold() for needle in needles):
            return points
    return None


MISSING_SUMMARY_PROFILES = {
    "2024-12-17-rwa-plume-l2-l1": "Plume从L2转向L1，意在围绕原生RWA资产建立独立生态与价值捕获，RWAFi扩张和真实资产供给决定战略成效。",
    "2024-12-19-cat": "西蒙猫CAT依靠成熟IP和庞大社交影响力获得初始关注，锁仓、回购销毁与真实品牌转化共同决定代币长期价值。",
    "2024-12-20-compute-decentralized-depin-spheron-ai": "Spheron通过分层节点降低去中心化GPU供给门槛，资源规格和网络扩展性适配AI训练，但持续需求仍是商业化关键。",
    "2024-12-20-pengu": "PENGU以胖企鹅IP影响力获得高估值起点，但NFT市场仍处低谷，品牌流量能否转化为持续需求决定价格支撑。",
    "2024-12-28-binance-18175601282290": "一季度潜在TGE项目普遍拥有高融资估值和基础设施定位，上线可能放大市场波动，收入能力与流通结构决定承接空间。",
    "2025-01-08-depin-multiple": "Multiple聚焦网络数据传输并以节点插件扩展资源供给，低门槛部署有利于规模增长，但实际传输需求仍需持续验证。",
    "2025-01-21-3": "Lista等低市值项目依靠BNB Chain生态地位、收入增长与流动性服务形成重估线索，但兑现仍取决于链上需求扩张。",
    "2025-02-01-aethir-gpu-nodefi": "Aethir把GPU供给、节点质押与NodeFi结合，试图构建更灵活的算力市场，增长关键在于降低供应商准入与资金占用。",
    "2025-02-04-odin-fun": "Odin Fun将链抽象能力用于资产发行，相比UniversalX更偏向发行场景，创新价值取决于跨链体验和用户需求能否形成规模。",
    "2025-02-05-depin-bless-nnapp": "Bless以共享计算节点为nnAPP提供随用户增长扩展的算力，低门槛节点网络带来冷启动优势，也需要真实应用消化供给。",
    "2025-02-06-megaeth": "MegaETH凭借高性能执行与以太坊核心生态支持获得高估值预期，但融资定价、解锁结构和真实应用将决定上线表现。",
    "2025-02-08-solana-pump-fun2-0-pumpkin-fun": "Pumpkin fun通过手续费、代币回购和开发者激励改造发币平台模式，持续价值取决于交易活跃度与机制能否约束短期抛售。",
    "2025-02-12-pi-network": "Pi Network拥有庞大社区与节点基础，并采用改进的联邦拜占庭共识，但代币分布、真实流通和生态使用决定网络价值。",
    "2025-02-20-atom": "双币投资需要同时承担价格波动与执行风险，而ATOM更深层的问题在于生态增长未能有效回流代币，价值捕获仍然不足。",
    "2025-02-26-solana-time-fun": "Time fun把创作者时间代币化并采用动态定价，模式比Pump fun更复杂，能否吸引真实付费需求决定其扩张上限。",
    "2025-02-28-roam-tge": "ROAM以积分燃烧和动态兑换比例连接长期贡献与代币释放，机制有助于缓解抛压，但仍需平衡激励和真实网络使用。",
    "2025-03-07-gps": "GoPlus把风险识别、安全交易和实时拦截整合为链上安全层，覆盖场景广度与持续调用需求决定GPS的价值捕获。",
    "2025-03-14-zksync-20": "ZKsync的RWA规模增长主要由Tradable数据推动，但链上资金证明与公开信息存在差异，资产真实性仍需进一步核验。",
    "2025-03-22-launchpool-nillion": "Nillion以盲计算在加密数据上执行运算，为AI和隐私场景提供基础设施，贡献型网络和实际计算需求决定长期采用。",
    "2025-01-01-20": "Web3仍有一批产品通过预测市场、基础设施和真实服务创造收入，判断价值应回到用户需求、收入效率与竞争壁垒。",
    "2025-01-06-pow-4": "PoW基础设施依靠外部资源投入和原生分发形成独立价值，扩容、算力与网络结构决定项目能否脱离短期市场叙事。",
    "2025-01-12-ai": "AI抽象层与链抽象正在共同降低资产和应用交互门槛，产品差异最终取决于数据整合、跨链能力与真实使用体验。",
    "2025-01-22-booster-rollup": "Booster Rollup试图减少EVM应用跨L2重复部署成本，但以太坊扩容层之间的竞争与流动性分散仍限制其价值空间。",
    "2025-01-29-desci": "DeSci覆盖科研基础设施、数据和研究融资，但成果转化周期长、失败率高，商业闭环仍比短期市场热度更重要。",
    "2025-02-23-12-tge": "链抽象协议正围绕通用账户、跨链交互和AI代理降低使用门槛，竞争关键在于产品落地而非单纯TGE预期。",
    "2025-03-15-binance-21576651114778": "比特币减半持续压缩矿工收入，安全预算需要可持续费用市场支撑，BTCFi仍缺少垄断性协议并保留较大增长空间。",
    "2025-03-27-rwa": "证券代币化规模短期快速增长，但部分资产缺少链上资金证明，RWA扩张必须同时验证发行真实性与可追溯性。",
    "2024-12-18-binance-17718746563425": "币安钱包和上币体系正强化链上数据、用户热度与基本面筛选，并通过Alpha和空投工具降低用户进入Web3的门槛。",
    "2024-12-29-dappos-30": "dappOS通过求解器和意图资产自动完成跨链操作，节点清算机制保障执行，核心价值在于降低复杂交互成本。",
    "2024-12-30-solana-l2": "Solana自身性能较强，使L2扩容必要性受到质疑；应用链若要成立，必须证明独立代币经济和流动性价值。",
    "2025-01-13-binance-18845998934201": "内容平台账号增长依赖主题垂直度、数据反馈和平台激励，持续产出与用户互动比短期奖励更能形成长期影响力。",
    "2025-01-24-binance-19355006124497": "UniversalX是Particle链抽象基础设施的应用入口，两者通过后台协调流动性，把跨链操作转化为更直接的用户体验。",
    "2025-02-03-yzi-labs": "YZi Labs投资与币安上币均采用多维筛选和集体决策，二级市场定价应减少对一级高估值的被动承接。",
    "2025-02-28-24-myshell-binance": "MyShell通过AI应用市场、份额定价和创作者优先权构建AI创作者经济，供需与应用质量决定资产价值。",
    "2025-03-18-binance-21702982030618": "双代币与积分加代币模式都在分离治理和使用价值，区别在兑换机制、激励周期以及对流动性和用户理解成本的影响。",
    "2025-03-22-binance-21890984752466": "空投从无条件追溯转向积分和贡献证明，有助于冷启动与留存，但提前承诺也可能放大刷量和融资导向。",
}


def generate_missing_summary(path: Path, title: str, body: str, points: list[str]) -> dict:
    summary = MISSING_SUMMARY_PROFILES.get(path.stem, "")
    if not summary:
        raise ValueError("缺少经审定的摘要配置，拒绝使用通用模板")
    summary = ensure_terminal_punctuation(summary)
    if not points:
        raise ValueError("缺少核心观点，无法生成description")
    description = ensure_terminal_punctuation(
        clean_generated_text(summary.rstrip("。") + "。" + points[0] + (points[1] if len(points) > 1 else ""))
    )
    if visible_length(description) > 160:
        description = ensure_terminal_punctuation(clip_sentence(description, 155))
    quality = summary_copy_quality(summary, description, title, body)
    if not quality["passed"]:
        raise ValueError("摘要或描述未通过质量校验：" + json.dumps(quality, ensure_ascii=False))
    return {"summary": summary, "description": description, "quality": quality}


def generate_core_points(title: str, category: str, body: str) -> dict:
    """Generate independently validated core points without rewriting article prose."""
    blocks = markdown_body_blocks(body)
    if not blocks:
        raise ValueError("正文为空，无法生成核心观点")

    attempts: list[dict] = []
    profile = core_points_profile(title)
    if profile:
        quality = core_points_quality(profile, title, body)
        attempts.append({"attempt": "domain-profile", **quality})
        if quality["passed"]:
            return {"key_points": quality["points"], "quality": quality, "attempts": attempts}

    candidates: list[tuple[int, int, str]] = []
    for order, sentence in enumerate(sentences(blocks)):
        point = ensure_terminal_punctuation(clean_generated_text(sentence))
        if not 18 <= visible_length(point) <= 140:
            continue
        if contains_banned_language(point) or has_abnormal_punctuation(point):
            continue
        if any(
            phrase in point
            for phrase in (
                "本篇", "本文", "这篇文章", "这篇帖子", "通篇", "我个人", "我知道",
                "我认为", "我觉得", "大家", "朋友", "下面", "上图", "下图", "原推",
                "介绍完", "聊聊", "先说", "先看", "再看", "接下来", "欢迎讨论",
            )
        ):
            continue
        if "？" in point or "?" in point:
            continue
        if re.match(r"^\s*(?:\d+[.、）)]|[（(]\d+[）)])", point):
            continue
        if point.rstrip("。！").endswith(("假如", "比如", "因为", "但是", "不过", "两个", "如下")):
            continue
        if text_similarity(point, title) >= 0.82:
            continue
        candidates.append((-sentence_score(point, title), order, point))

    selected: list[str] = []
    for _score, _order, point in sorted(candidates):
        if all(text_similarity(point, prior) < 0.8 for prior in selected):
            selected.append(point)
        if len(selected) == 5:
            break
    if len(selected) > 3:
        selected = selected[:4]
    quality = core_points_quality(selected, title, body)
    attempts.append({"attempt": "evidence-extraction", **quality})
    if quality["passed"]:
        return {"key_points": quality["points"], "quality": quality, "attempts": attempts}
    raise ValueError("核心观点生成未通过质量校验：" + json.dumps(attempts, ensure_ascii=False))


CARD_THEME_TERMS = (
    ("用户需求", ("用户需求", "真实需求", "用户增长", "用户留存", "用户")),
    ("产品体验", ("产品体验", "使用体验", "交互", "应用")),
    ("代币激励", ("代币激励", "通证经济", "代币经济", "积分")),
    ("价值捕获", ("价值捕获", "价值", "收入", "利润", "收益")),
    ("流动性", ("流动性", "交易量", "资金")),
    ("供需关系", ("供需", "供应", "需求")),
    ("网络规模", ("网络规模", "节点", "设备数量", "覆盖")),
    ("生态扩张", ("生态", "开发者", "合作伙伴", "扩张")),
    ("安全与合规", ("安全", "合规", "监管", "风险")),
    ("成本效率", ("成本", "效率", "费用", "毛利率")),
    ("技术架构", ("架构", "性能", "Layer", "L1", "L2", "基础设施")),
    ("商业模式", ("商业模式", "业务模式", "盈利模式")),
    ("市场竞争", ("竞争", "市场份额", "护城河", "差异化")),
    ("资产机制", ("RWA", "资产", "质押", "借贷")),
)


def infer_card_subject(title: str, body: str) -> str:
    special = (
        ("PMF", "Web3产品PMF"),
        ("NVDA", "英伟达"),
        ("英伟达", "英伟达"),
        ("BIO", "BIO Protocol"),
        ("Roam", "Roam"),
        ("Plume", "Plume"),
        ("Aethir", "Aethir"),
        ("Odin Fun", "Odin Fun"),
        ("Solv", "Solv"),
        ("dappOS", "dappOS"),
        ("老虎证券", "老虎证券"),
    )
    haystack = f"{title}\n{body}"
    for needle, subject in special:
        if needle.casefold() in haystack.casefold():
            return subject
    entities = [
        item
        for item in re.findall(r"[A-Za-z][A-Za-z0-9.$-]{1,24}", title)
        if item.casefold() not in {"web3", "ai", "gpu", "rwa", "dex", "tge", "l1", "l2"}
    ]
    if entities:
        return entities[0]
    cleaned = re.sub(
        r"^(解析|解读|聊聊|关于|为什么|如何看|如何|什么是|一览|盘点|直面|市场上)",
        "",
        title,
    )
    cleaned = re.split(r"[丨｜：:，,。！？?!（(\[【]", cleaned, maxsplit=1)[0]
    cleaned = cleaned.strip("[]【】（）() ~～")
    if 2 <= visible_length(cleaned) <= 18:
        return cleaned
    return "该研究对象"


def detected_card_themes(body: str) -> list[str]:
    themes: list[str] = []
    for label, needles in CARD_THEME_TERMS:
        if any(needle.casefold() in body.casefold() for needle in needles):
            themes.append(label)
    return themes


def card_copy_profile(title: str) -> dict | None:
    """High-confidence domain profiles used before the conservative generic synthesizer."""
    profiles = (
        (
            ("英伟达", "NVDA"),
            "英伟达增长虽有所放缓，但Blackwell放量与推理模型算力需求仍支撑高端GPU市场，毛利率压力更接近新品爬坡影响。",
            "英伟达最新财报显示数据中心仍是增长核心，Blackwell已经贡献可观收入，但整体增速与毛利率出现回落。结合DeepSeek效率提升和推理模型算力消耗，GPU需求尚未进入平台期，当前盈利压力主要来自新品产能爬坡与成本波动。",
        ),
        (
            ("游戏驿站", "GameStop"),
            "游戏驿站季度利润创下纪录，但主要增量来自衍生品纸面收益，主营业务仍待修复，现金储备如何支持转型才是重估关键。",
            "游戏驿站本季营收和净利润超预期，但近七成利润来自eBay收购相关衍生品的未实现收益，传统硬件与软件业务仍在收缩。收藏品增长、充足流动性、股份回购与潜在收购共同决定公司能否改善主营盈利能力。",
        ),
        (
            ("DePIN丨Roam",),
            "Roam以免费WiFi、硬件节点和代币激励连接无线网络供需，设备规模领先，但长期价值仍取决于补贴退坡后的自给能力。",
            "Roam通过硬件节点、WiFi地图、eSIM和代币激励构建全球无线网络，设备数量在DePIN项目中居前。其供需循环依赖免费连接和用户激励，网络扩张能否逐步摆脱补贴并形成持续的资源供给与运营收益，是长期价值的关键。",
        ),
        (
            ("BIO上线", "BIO Protocol"),
            "BIO Protocol通过科研项目与知识产权代币化连接资金和科学家，机会在DeSci融资效率，持续性则取决于成果转化与治理。",
            "BIO Protocol面向去中心化科学场景，为科研团队提供社区、融资、治理和知识产权代币化工具。BioDAO机制试图改善科研融资和资源共享，但项目成果转化、社区治理效率以及真实科研需求能否持续进入网络，仍是价值兑现的核心条件。",
        ),
        (
            ("以太 $ETH", "以太坊价值"),
            "以太坊高费用反映区块空间需求，也削弱普通用户体验；长期价值取决于主网安全、Layer2扩容与生态价值捕获能否平衡。",
            "以太坊交易费用高昂既说明区块空间仍有真实需求，也暴露了主网使用门槛。安全性、去中心化、Layer2扩容和生态网络效应共同构成ETH的价值来源，扩容过程还需要维持主网价值捕获与长期安全预算。",
        ),
        (
            ("PMF",),
            "Web3产品是否找到PMF，关键不只在用户增长，而在真实需求、留存、付费意愿与持续使用能否同时成立。",
            "Web3产品的PMF不能只靠补贴、空投或短期活跃判断。以链抽象和UniversalX为例，可以从真实需求、解决方案有效性、用户持续使用、留存与付费意愿反向验证产品市场匹配，并区分依赖外部输血的增长和可持续需求。",
        ),
        (
            ("Coinbase分发权", "入口垄断"),
            "Base的长期壁垒不只来自L2性能，而是Coinbase用户入口、稳定币流动性和应用分发共同形成的链上转化能力。",
            "Base在L2竞争中的优势并非单纯来自性能或费用，而在于Coinbase能够把交易所用户、账户体系、稳定币流动性和应用入口导向链上。分发效率与生态转化形成的入口控制，可能比技术参数构成更难复制的护城河。",
        ),
    )
    for needles, summary, description in profiles:
        if any(needle.casefold() in title.casefold() for needle in needles):
            return {"summary": summary, "description": description}
    return None


def generate_card_copy(title: str, category: str, body: str) -> dict:
    blocks = markdown_body_blocks(body)
    body_text = " ".join(blocks)
    subject = infer_card_subject(title, body_text)
    themes = detected_card_themes(body_text)
    if len(themes) < 3:
        raise ValueError("正文中可验证的主题不足3项，无法安全生成独立摘要")
    themes = themes[:5]
    attempts: list[dict] = []
    profile = card_copy_profile(title)
    if profile:
        profile_quality = summary_copy_quality(
            profile["summary"], profile["description"], title, body
        )
        attempts.append({"attempt": "domain-profile", **profile_quality})
        if profile_quality["passed"]:
            return {
                "summary": profile["summary"],
                "description": profile["description"],
                "quality": profile_quality,
                "attempts": attempts,
            }
    variants = (
        (
            f"{subject}的核心逻辑取决于{themes[0]}能否形成稳定机制，{themes[1]}影响扩张效率，而{themes[2]}将决定长期价值兑现。",
            f"{subject}的研究重点包括{themes[0]}、{themes[1]}与{themes[2]}之间的关系。相关判断还需要结合{themes[-1]}观察其运行机制是否能够持续，并据此识别增长空间、价值来源以及可能限制后续发展的关键风险。",
        ),
        (
            f"{subject}正围绕{themes[0]}建立差异化，{themes[1]}提供增长基础，{themes[2]}与{themes[-1]}共同决定这一模式能否持续扩张。",
            f"围绕{subject}的实际进展，核心问题在于{themes[0]}能否转化为稳定的{themes[1]}优势，同时由{themes[2]}验证增长质量。{themes[-1]}既构成后续扩张条件，也界定了这一模式的适用范围与潜在风险。",
        ),
    )
    for attempt, (summary, description) in enumerate(variants, start=1):
        summary = ensure_terminal_punctuation(summary)
        description = ensure_terminal_punctuation(description)
        quality = summary_copy_quality(summary, description, title, body)
        attempts.append({"attempt": attempt, **quality})
        if quality["passed"]:
            return {
                "summary": summary,
                "description": description,
                "quality": quality,
                "attempts": attempts,
            }
    raise ValueError("两次独立摘要生成均未通过质量校验：" + json.dumps(attempts, ensure_ascii=False))


def resolve_html(field: str, html_dir: Path) -> Path | None:
    field = clean_text(field)
    if not field:
        return None
    candidate = html_dir / field
    if candidate.suffix.lower() != ".html":
        candidate = html_dir / f"{field}.html"
    return candidate if candidate.is_file() else None


@dataclass
class MatchResult:
    article: ExistingArticle | None
    method: str
    error: str = ""


def match_existing_article(
    values: dict,
    html_path: Path | None,
    by_id: dict[str, list[ExistingArticle]],
    by_url: dict[str, list[ExistingArticle]],
    by_title: dict[str, list[ExistingArticle]],
) -> MatchResult:
    content_id = clean_text(values.get("ID"))
    source_url = normalize_url(values.get("Binance链接"))
    if content_id:
        id_matches = by_id.get(content_id, [])
        if len(id_matches) > 1:
            return MatchResult(None, "ID=content_id", f"content_id匹配到多个Markdown：{content_id}")
        if len(id_matches) == 1:
            article = id_matches[0]
            if source_url and article.source_url and source_url != article.source_url:
                return MatchResult(
                    None,
                    "ID=content_id",
                    f"content_id匹配但source_url冲突：Excel={source_url}，Markdown={article.source_url}",
                )
            return MatchResult(article, "ID=content_id")

    if source_url:
        url_matches = by_url.get(source_url, [])
        if len(url_matches) > 1:
            return MatchResult(None, "Binance链接=source_url", f"source_url匹配到多个Markdown：{source_url}")
        if len(url_matches) == 1:
            article = url_matches[0]
            if article.content_id and article.content_id != content_id:
                return MatchResult(
                    None,
                    "Binance链接=source_url",
                    f"source_url匹配但content_id冲突：Excel={content_id}，Markdown={article.content_id}",
                )
            return MatchResult(article, "Binance链接=source_url")

    title_keys = {normalize_title(values.get("HTML文件名"))}
    if html_path:
        title_keys.add(normalize_title(html_path.stem))
        try:
            document = html_path.read_text(encoding="utf-8")
            title_keys.add(normalize_title(meta(document, "og:title") or meta(document, "twitter:title")))
        except (OSError, UnicodeDecodeError):
            pass
    title_keys.discard("")
    title_matches: dict[Path, ExistingArticle] = {}
    for title_key in title_keys:
        for article in by_title.get(title_key, []):
            title_matches[article.path] = article
    if len(title_matches) > 1:
        paths = "、".join(str(path) for path in sorted(title_matches, key=str))
        return MatchResult(None, "严格标准化标题", f"严格标准化标题匹配到多个Markdown：{paths}")
    if len(title_matches) == 1:
        article = next(iter(title_matches.values()))
        if article.content_id and article.content_id != content_id:
            return MatchResult(
                None,
                "严格标准化标题",
                f"标题严格匹配但content_id冲突：Excel={content_id}，Markdown={article.content_id}",
            )
        if source_url and article.source_url and source_url != article.source_url:
            return MatchResult(
                None,
                "严格标准化标题",
                f"标题严格匹配但source_url冲突：Excel={source_url}，Markdown={article.source_url}",
            )
        return MatchResult(article, "HTML文件名/HTML标题=Markdown标题（严格标准化）")
    return MatchResult(None, "未匹配")


def safe_write_binary(path: Path, data: bytes):
    if path.exists():
        if path.read_bytes() != data:
            raise FileExistsError(f"目标图片已存在且内容不同：{path}")
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def overwrite_binary_atomic(path: Path, data: bytes):
    """Atomically replace an existing binary for explicit maintenance modes."""
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(prefix=path.name + ".", dir=path.parent)
    os.close(handle)
    temporary_path = Path(temporary_name)
    try:
        temporary_path.write_bytes(data)
        os.replace(temporary_path, path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def mirror_file(source: Path, destination: Path):
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        if destination.read_bytes() != source.read_bytes():
            raise FileExistsError(f"镜像图片已存在且内容不同：{destination}")
        return
    shutil.copy2(source, destination)


def build_new_article(
    html_path: Path,
    content_id: str,
    excel_url: str,
    category_label: str,
    featured: bool,
    status: str,
    site_root: Path,
    dry_run: bool,
    allow_remote_images: bool = False,
) -> dict:
    document = html_path.read_text(encoding="utf-8")
    title = clean_text(meta(document, "og:title") or meta(document, "twitter:title"))
    if not title:
        raise ValueError("HTML 中找不到原始标题（og:title/twitter:title）")
    html_url = normalize_url(meta(document, "og:url"))
    excel_url = normalize_url(excel_url)
    if excel_url and html_url and excel_url != html_url:
        raise ValueError(f"Excel Binance链接与 HTML og:url 不一致：{excel_url} != {html_url}")
    source_url = excel_url or html_url
    if not source_url:
        raise ValueError("HTML 与 Excel 均缺少 Binance 原链接")
    published = published_at(document)
    if not published:
        raise ValueError("HTML 中找不到原始发布时间；禁止用当前日期或文件时间代替")

    article_html, article_start = find_article_html(document)
    parser = ArticleParser()
    parser.feed(article_html)
    parser.finish()
    body_blocks = plain_body_blocks(parser)
    if len(body_blocks) < 2:
        raise ValueError("提取到的正文过短，拒绝导入")

    category = CATEGORY_MAP[category_label]
    classification = category_frontmatter(category_label)
    date_prefix = published.date().isoformat()
    content_dir = site_root / "content" / "zh-cn" / category["section"] / category["directory"]
    if not content_dir.is_dir():
        raise FileNotFoundError(f"现有内容目录不存在，拒绝新建重复栏目：{content_dir}")
    slug = ensure_safe_article_slug(
        f"{date_prefix}-{slugify(title, source_url)}",
        source_url,
        site_root,
        content_dir,
        category["section"],
        date_prefix,
    )
    article_path = content_dir / f"{slug}.md"
    if article_path.exists():
        raise FileExistsError(f"目标 Markdown 已存在，拒绝覆盖：{article_path}")

    cover = select_cover(
        document,
        article_start,
        title,
        category_label,
        html_path,
        allow_remote=allow_remote_images,
    )
    cover_data = cover.output_data
    cover_width, cover_height = cover.width, cover.height
    cover_ext = cover.output_format

    covers_dir = site_root / "assets" / "images" / "covers"
    posts_dir = site_root / "assets" / "images" / "posts"
    static_covers_dir = site_root / "static" / "images" / "covers"
    static_posts_dir = site_root / "static" / "images" / "posts"
    cover_name = (
        f"{slug}-cover-placeholder.png"
        if cover.is_placeholder
        else f"{slug}-cover.{cover_ext}"
    )
    cover_asset = covers_dir / cover_name
    cover_static = static_covers_dir / cover_name

    image_urls: list[str | None] = [None] * len(parser.images)
    image_plans: list[tuple[Path, Path, bytes]] = []
    body_image_source_types: list[str] = []
    missing_body_images = 0
    image_number = 0
    variables = css_variable_images(document)
    for index, image in enumerate(parser.images):
        attrs = {str(key): value or "" for key, value in image.items()}
        sources = attribute_image_sources(attrs, variables, html_path)
        marker_source = next((source for source, _ in sources), attrs.get("src", ""))
        if is_noise_image(attrs, marker_source):
            continue
        data = None
        extension = ""
        selected_kind = ""
        local_errors: list[str] = []
        for source, source_kind in sources:
            if cover.source and source == cover.source:
                data = None
                selected_kind = "与封面重复"
                break
            if source_kind == "remote" and not allow_remote_images:
                continue
            try:
                candidate_data = read_image_source(
                    source, source_kind, html_path, allow_remote=allow_remote_images
                )
                _, _, candidate_extension = image_info(candidate_data)
                data = candidate_data
                extension = candidate_extension
                selected_kind = source_kind
                break
            except Exception as exc:
                if source_kind != "remote":
                    local_errors.append(f"{source_kind}: {exc}")
        if selected_kind == "与封面重复":
            continue
        if data is None:
            if local_errors:
                raise ValueError(
                    f"正文第 {index + 1} 张内嵌图片损坏：{'；'.join(local_errors)}"
                )
            missing_body_images += 1
            continue
        image_number += 1
        name = f"{slug}-image-{image_number:02d}.{extension}"
        asset_path = posts_dir / name
        static_path = static_posts_dir / name
        image_plans.append((asset_path, static_path, data))
        body_image_source_types.append(selected_kind)
        image_urls[index] = f"/images/posts/{name}"

    body_markdown = render_markdown(parser, image_urls)
    try:
        seo_geo = generate_seo_geo(body_blocks, title, category_label, body_markdown)
    except Exception as exc:
        seo_geo = failed_seo_geo(str(exc))
    summary = seo_geo["summary"]
    description = seo_geo["description"]
    key_points = seo_geo["key_points"]
    seo_keywords = seo_geo["seo_keywords"]
    faq = seo_geo["faq"]

    lines = [
        "---",
        f"title: {yaml_string(title)}",
        f"content_id: {yaml_string(content_id)}",
        f"date: {yaml_string(published.isoformat(timespec='seconds'))}",
        f"slug: {yaml_string(slug)}",
        f"category: {yaml_string(classification['category'])}",
        f"subcategory: {yaml_string(classification['subcategory'])}",
        f"content_type: {yaml_string(classification['content_type'])}",
        f"category_key: {yaml_string(classification['category_key'])}",
        f"category_label: {yaml_string(classification['category_label'])}",
        "categories:",
        f"  - {yaml_string(category_label)}",
        f"featured: {str(featured).lower()}",
        f"status: {yaml_string(status)}",
        f"draft: {str(draft_for_status(status)).lower()}",
        f"author: {yaml_string(AUTHOR)}",
        'source: "Binance Square"',
        f"source_url: {yaml_string(source_url)}",
        f"cover: {yaml_string('/images/covers/' + cover_name)}",
        f"seo_geo_status: {yaml_string(seo_geo['_final_status'])}",
        f"summary: {yaml_string(summary)}",
        f"description: {yaml_string(description)}",
    ]
    if key_points:
        lines += ["key_points:", *[f"  - {yaml_string(point)}" for point in key_points]]
    else:
        lines.append("key_points: []")
    if seo_keywords:
        lines += [
            "seo_keywords:",
            *[f"  - {yaml_string(word)}" for word in seo_keywords],
            "tags:",
            *[f"  - {yaml_string(word)}" for word in seo_keywords],
        ]
    else:
        lines += ["seo_keywords: []", "tags: []"]
    if faq:
        lines.append("faq:")
        for item in faq:
            lines += [
                f"  - question: {yaml_string(item['question'])}",
                f"    answer: {yaml_string(item['answer'])}",
            ]
    else:
        lines.append("faq: []")
    structured_body = render_structured_article_body(body_markdown)
    lines += ["---", "", structured_body, ""]

    if not dry_run:
        safe_write_binary(cover_asset, cover_data)
        mirror_file(cover_asset, cover_static)
        for asset_path, static_path, data in image_plans:
            safe_write_binary(asset_path, data)
            mirror_file(asset_path, static_path)
        article_path.write_text("\n".join(lines), encoding="utf-8", newline="\n")

    return {
        "title": title,
        "content_id": content_id,
        "source_url": source_url,
        "status": status,
        "published_at": published.isoformat(timespec="seconds"),
        "article_path": article_path,
        "cover_asset": cover_asset,
        "cover_static": cover_static,
        "cover_dimensions": f"{cover_width}x{cover_height}",
        "cover_candidate_count": cover.candidate_count,
        "detected_image_resource_count": cover.resource_count,
        "detected_image_resource_types": cover.resource_types,
        "cover_selected_source": cover.source_label,
        "cover_source_type": cover.source_kind,
        "cover_html_region": cover.region,
        "cover_original_dimensions": f"{cover.width}x{cover.height}",
        "cover_is_true_top_visual": cover.is_true_top_visual,
        "cover_needs_normalization": cover.needs_normalization,
        "cover_output_dimensions": f"{cover.output_width}x{cover.output_height}",
        "cover_used_remote_request": cover.used_remote,
        "cover_used_placeholder": cover.is_placeholder,
        "cover_status": cover.cover_status,
        "cover_will_continue_import": True,
        "body_images": image_number,
        "body_image_source_types": body_image_source_types,
        "body_images_missing": missing_body_images,
        "image_paths": [str(item[0]) for item in image_plans],
        "summary_length": visible_length(summary),
        "description_length": visible_length(description),
        "key_points": len(key_points),
        "seo_keywords": len(seo_keywords),
        "faq": len(faq),
        "summary": summary,
        "description": description,
        "key_point_values": key_points,
        "seo_keyword_values": seo_keywords,
        "faq_values": faq,
        "quality_validation": seo_geo["_quality"],
        "generation_attempts": seo_geo["_attempts"],
        "seo_geo_final_status": seo_geo["_final_status"],
        "seo_geo_error": seo_geo.get("_error", ""),
        "has_body_heading": "## 正文" in structured_body,
    }


@dataclass
class SyncReport:
    excel_total_rows: int = 0
    selected_rows: list[int] = field(default_factory=list)
    normalized_pending: int = 0
    normalized_featured_no: int = 0
    matched_old: int = 0
    old_category_updates: int = 0
    old_featured_updates: int = 0
    old_status_updates: int = 0
    imported_new: int = 0
    skipped: int = 0
    real_covers: int = 0
    placeholder_covers: int = 0
    seo_generated: int = 0
    seo_preserved: int = 0
    seo_completed: int = 0
    seo_processing_failed: int = 0
    processing_failed: int = 0
    missing_html: list[dict] = field(default_factory=list)
    unmatched_articles: list[dict] = field(default_factory=list)
    invalid_categories: list[dict] = field(default_factory=list)
    failures: list[dict] = field(default_factory=list)
    generated_paths: list[dict] = field(default_factory=list)
    details: list[dict] = field(default_factory=list)
    status_to_pending: list[dict] = field(default_factory=list)
    featured_to_no: list[dict] = field(default_factory=list)
    matched_to_published: list[dict] = field(default_factory=list)
    seo_status_changes: list[dict] = field(default_factory=list)
    seo_failures: list[dict] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "Excel总行数": self.excel_total_rows,
            "本次处理行": self.selected_rows,
            "标准化为待导入的行数": self.normalized_pending,
            "标准化为精选否的行数": self.normalized_featured_no,
            "成功匹配旧文章数量": self.matched_old,
            "旧文章分类调整数量": self.old_category_updates,
            "旧文章精选调整数量": self.old_featured_updates,
            "旧文章状态调整数量": self.old_status_updates,
            "文章导入成功数量": self.imported_new,
            "新文章导入数量": self.imported_new,
            "新增文章数量": self.imported_new,
            "更新文章数量": self.matched_old,
            "跳过数量": self.skipped,
            "错误数量": self.processing_failed,
            "使用真实封面数量": self.real_covers,
            "使用占位封面数量": self.placeholder_covers,
            "SEO/GEO已完成数量": self.seo_completed,
            "SEO/GEO处理失败数量": self.seo_processing_failed,
            "SEO/GEO字段生成数量": self.seo_generated,
            "SEO/GEO保留数量": self.seo_preserved,
            "整篇文章处理失败数量": self.processing_failed,
            "处理失败数量": self.processing_failed,
            "找不到HTML的行": self.missing_html,
            "无法匹配文章列表": self.unmatched_articles,
            "无效分类列表": self.invalid_categories,
            "失败原因": self.failures,
            "每篇生成路径": self.generated_paths,
            "状态将改为待导入": self.status_to_pending,
            "精选将改为否": self.featured_to_no,
            "旧文匹配后将改为已发布": self.matched_to_published,
            "SEO/GEO状态将补充或调整": self.seo_status_changes,
            "SEO/GEO失败原因": self.seo_failures,
            "处理明细": self.details,
        }


def row_ref(row_number: int, values: dict) -> dict:
    return {"Excel行": row_number, "ID": values.get("ID"), "HTML文件名": values.get("HTML文件名")}


def parse_rows(value: str, maximum: int) -> list[int]:
    rows = []
    for item in value.split(","):
        item = item.strip()
        if not item:
            continue
        if not item.isdigit():
            raise ValueError(f"无效 Excel 行号：{item}")
        number = int(item)
        if number < 2 or number > maximum:
            raise ValueError(f"Excel 行号超出范围：{number}（有效范围 2-{maximum}）")
        if number not in rows:
            rows.append(number)
    if not rows:
        raise ValueError("--rows 不能为空")
    return rows


def save_workbook_atomic(workbook, path: Path):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=path.parent, delete=False) as handle:
        temp_path = Path(handle.name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def ensure_seo_status_column(sheet, headers: dict[str, int], dry_run: bool) -> tuple[int, bool]:
    if SEO_STATUS_COLUMN in headers:
        return headers[SEO_STATUS_COLUMN], False
    column = sheet.max_column + 1
    headers[SEO_STATUS_COLUMN] = column
    if dry_run:
        return column, True
    source_column = max(1, column - 1)
    sheet.cell(1, column).value = SEO_STATUS_COLUMN
    sheet.cell(1, column)._style = copy(sheet.cell(1, source_column)._style)
    sheet.cell(1, column).font = copy(sheet.cell(1, source_column).font)
    sheet.cell(1, column).fill = copy(sheet.cell(1, source_column).fill)
    sheet.cell(1, column).alignment = copy(sheet.cell(1, source_column).alignment)
    sheet.cell(1, column).border = copy(sheet.cell(1, source_column).border)
    sheet.column_dimensions[get_column_letter(column)].width = 16
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row_number, column)._style = copy(sheet.cell(row_number, source_column)._style)
    for table in sheet.tables.values():
        table.ref = f"A1:{get_column_letter(column)}{sheet.max_row}"
    return column, True


def worksheet_rows(sheet, headers: dict[str, int]) -> list[int]:
    return [
        row
        for row in range(2, sheet.max_row + 1)
        if any(sheet.cell(row, headers[column]).value not in (None, "") for column in REQUIRED_COLUMNS)
    ]


def worksheet_values(sheet, headers: dict[str, int], row_number: int) -> dict:
    columns = [*REQUIRED_COLUMNS, SEO_STATUS_COLUMN]
    result = {}
    for column in columns:
        value = sheet.cell(row_number, headers[column]).value
        result[column] = clean_text(str(value)) if value is not None else ""
    return result


class PreflightValidationError(ValueError):
    def __init__(self, errors: list[dict]):
        self.errors = errors
        super().__init__(f"同步前Excel检查失败，共 {len(errors)} 项错误")


def _duplicate_field_errors(
    rows: list[tuple[int, dict]],
    field_name: str,
    error_type: str,
    normalizer,
) -> list[dict]:
    grouped: dict[str, list[int]] = {}
    display_values: dict[str, str] = {}
    for row_number, values in rows:
        raw = clean_text(values.get(field_name))
        if not raw:
            continue
        key = normalizer(raw)
        if not key:
            continue
        grouped.setdefault(key, []).append(row_number)
        display_values.setdefault(key, raw)
    return [
        {"类型": error_type, "字段": field_name, "值": display_values[key], "Excel行": numbers}
        for key, numbers in grouped.items()
        if len(numbers) > 1
    ]


def preflight_validate(
    sheet,
    headers: dict[str, int],
    data_rows: list[int],
    html_dir: Path,
    by_id: dict[str, list[ExistingArticle]],
    by_url: dict[str, list[ExistingArticle]],
    by_title: dict[str, list[ExistingArticle]],
) -> None:
    """Fail atomically before any row is synchronized when control data is unsafe."""
    rows = [(row_number, worksheet_values(sheet, headers, row_number)) for row_number in data_rows]
    errors: list[dict] = []

    for row_number, values in rows:
        if not values["ID"]:
            errors.append({"类型": "ID为空", **row_ref(row_number, values)})
        if not values["HTML文件名"]:
            errors.append({"类型": "空标题", **row_ref(row_number, values)})
        if not values["目标分类"]:
            errors.append({"类型": "空分类", **row_ref(row_number, values)})
        elif values["目标分类"] not in CATEGORY_MAP:
            errors.append(
                {"类型": "无效分类", **row_ref(row_number, values), "目标分类": values["目标分类"]}
            )
        if values["状态"] and values["状态"] not in ALLOWED_STATUSES:
            errors.append(
                {"类型": "无效状态", **row_ref(row_number, values), "状态": values["状态"] or "（空）"}
            )

        html_path = resolve_html(values["HTML文件名"], html_dir)
        if html_path:
            try:
                document = html_path.read_text(encoding="utf-8")
                html_title = clean_text(meta(document, "og:title") or meta(document, "twitter:title"))
            except (OSError, UnicodeDecodeError):
                html_title = ""
            if not html_title:
                errors.append({"类型": "HTML空标题", **row_ref(row_number, values)})

    errors += _duplicate_field_errors(rows, "ID", "ID重复", lambda value: value.casefold())
    errors += _duplicate_field_errors(rows, "Binance链接", "Binance链接重复", normalize_url)
    errors += _duplicate_field_errors(
        rows,
        "HTML文件名",
        "HTML文件名重复",
        lambda value: re.sub(r"\.html?$", "", unicodedata.normalize("NFKC", value), flags=re.I).casefold(),
    )

    for content_id, articles in by_id.items():
        if len(articles) > 1:
            errors.append(
                {
                    "类型": "Markdown content_id重复",
                    "content_id": content_id,
                    "Markdown": [str(article.path) for article in articles],
                }
            )
    for source_url, articles in by_url.items():
        if len(articles) > 1:
            errors.append(
                {
                    "类型": "Markdown source_url重复",
                    "source_url": source_url,
                    "Markdown": [str(article.path) for article in articles],
                }
            )

    # Detect cross-key conflicts before an update can attach an ID to the wrong article.
    for row_number, values in rows:
        if not values["ID"]:
            continue
        html_path = resolve_html(values["HTML文件名"], html_dir)
        match = match_existing_article(values, html_path, by_id, by_url, by_title)
        if match.error and ("冲突" in match.error or "多个Markdown" in match.error):
            errors.append({"类型": "匹配键冲突", **row_ref(row_number, values), "原因": match.error})

    if errors:
        raise PreflightValidationError(errors)


def run_normalize_excel(
    args,
    workbook,
    sheet,
    headers: dict[str, int],
    data_rows: list[int],
    by_id: dict[str, list[ExistingArticle]],
    by_url: dict[str, list[ExistingArticle]],
    by_title: dict[str, list[ExistingArticle]],
    excel_path: Path,
) -> SyncReport:
    report = SyncReport(excel_total_rows=len(data_rows), selected_rows=data_rows)
    workbook_changed = False
    for row_number in data_rows:
        values = worksheet_values(sheet, headers, row_number)
        html_path = resolve_html(values["HTML文件名"], args.html_dir_path)
        match = match_existing_article(values, html_path, by_id, by_url, by_title)
        article = match.article if not match.error else None
        old_status = values["状态"]
        new_status = normalize_status(old_status)
        old_featured = values["是否精选"]
        new_featured, _featured_bool = normalize_featured(old_featured)
        old_seo_status = values[SEO_STATUS_COLUMN]
        new_seo_status = normalize_seo_status(old_seo_status, article)

        if values["目标分类"] not in CATEGORY_MAP:
            report.invalid_categories.append(
                {**row_ref(row_number, values), "目标分类": values["目标分类"] or "（空）"}
            )
        if match.error:
            report.unmatched_articles.append({**row_ref(row_number, values), "原因": match.error})
        if article:
            report.matched_old += 1
            if new_status == "已发布":
                report.matched_to_published.append(
                    {
                        **row_ref(row_number, values),
                        "匹配方式": match.method,
                        "原值": old_status or "（空）",
                        "新值": new_status,
                        "动作": "保持" if old_status == new_status else "修改",
                    }
                )

        changes = {}
        if old_status != new_status:
            changes["状态"] = {"原值": old_status or "（空）", "新值": new_status}
            if new_status == "待导入":
                report.normalized_pending += 1
                report.status_to_pending.append({**row_ref(row_number, values), **changes["状态"]})
            if not args.dry_run:
                sheet.cell(row_number, headers["状态"]).value = new_status
                workbook_changed = True
        if old_featured != new_featured:
            changes["是否精选"] = {"原值": old_featured or "（空）", "新值": new_featured}
            if new_featured == "否":
                report.normalized_featured_no += 1
                report.featured_to_no.append({**row_ref(row_number, values), **changes["是否精选"]})
            if not args.dry_run:
                sheet.cell(row_number, headers["是否精选"]).value = new_featured
                workbook_changed = True
        if old_seo_status != new_seo_status:
            changes[SEO_STATUS_COLUMN] = {"原值": old_seo_status or "（空）", "新值": new_seo_status}
            report.seo_status_changes.append(
                {**row_ref(row_number, values), "匹配方式": match.method, **changes[SEO_STATUS_COLUMN]}
            )
            if not args.dry_run:
                sheet.cell(row_number, headers[SEO_STATUS_COLUMN]).value = new_seo_status
                workbook_changed = True
        if changes:
            report.details.append(
                {
                    **row_ref(row_number, values),
                    "匹配方式": match.method,
                    "Markdown": str(article.path) if article else "",
                    "标准化变更": changes,
                }
            )
    if workbook_changed and not args.dry_run:
        save_workbook_atomic(workbook, excel_path)
    return report


def run_sync(args) -> SyncReport:
    site_root = args.site_root.resolve()
    excel_path = (site_root / args.excel).resolve()
    html_dir = (site_root / args.html_dir).resolve()
    if not excel_path.is_file():
        raise FileNotFoundError(excel_path)
    if not html_dir.is_dir():
        raise FileNotFoundError(html_dir)
    args.html_dir_path = html_dir

    workbook = load_workbook(excel_path)
    sheet = workbook.active
    headers = {clean_text(str(cell.value or "")): cell.column for cell in sheet[1]}
    missing_headers = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing_headers:
        raise ValueError(f"Excel 缺少字段：{', '.join(missing_headers)}")
    _seo_column, seo_column_added = ensure_seo_status_column(sheet, headers, args.dry_run)
    data_rows = worksheet_rows(sheet, headers)
    by_id, by_url, by_title = content_index(site_root / "content" / "zh-cn")
    preflight_validate(sheet, headers, data_rows, html_dir, by_id, by_url, by_title)
    if args.normalize_excel:
        report = run_normalize_excel(
            args, workbook, sheet, headers, data_rows, by_id, by_url, by_title, excel_path
        )
        if seo_column_added:
            report.details.insert(
                0,
                {"结果": "dry-run将新增列" if args.dry_run else "已新增列", "字段": SEO_STATUS_COLUMN},
            )
            if not args.dry_run:
                save_workbook_atomic(workbook, excel_path)
        return report

    selected_rows = data_rows if args.all else parse_rows(args.rows, sheet.max_row)
    report = SyncReport(excel_total_rows=len(data_rows), selected_rows=selected_rows)
    workbook_changed = False
    planned_article_paths: set[Path] = set()

    def set_excel(row_number: int, column: str, value: str):
        nonlocal workbook_changed
        if sheet.cell(row_number, headers[column]).value != value:
            sheet.cell(row_number, headers[column]).value = value
            workbook_changed = True

    def mark_failed(row_number: int, values: dict, reason: str, bucket: str | None = None):
        item = {**row_ref(row_number, values), "原因": reason}
        report.failures.append(item)
        report.processing_failed += 1
        if bucket:
            getattr(report, bucket).append(item)
        report.details.append({**item, "结果": "处理失败", "SEO/GEO最终状态": "处理失败"})

    for row_number in selected_rows:
        values = worksheet_values(sheet, headers, row_number)
        category_label = values["目标分类"]
        source_url = normalize_url(values["Binance链接"])

        if category_label not in CATEGORY_MAP:
            mark_failed(row_number, values, f"无效目标分类：{category_label or '（空）'}", "invalid_categories")
            continue
        featured_text, featured = normalize_featured(values["是否精选"])
        html_path = resolve_html(values["HTML文件名"], html_dir)
        match = match_existing_article(values, html_path, by_id, by_url, by_title)
        if match.error:
            mark_failed(row_number, values, match.error, "unmatched_articles")
            continue
        existing = match.article
        desired_status = normalize_status(values["状态"])
        seo_status = normalize_seo_status(values[SEO_STATUS_COLUMN], existing)

        if existing:
            category = CATEGORY_MAP[category_label]
            classification = category_frontmatter(category_label)
            old_category = scalar_from_frontmatter(existing.frontmatter, "category_label")
            if old_category is None:
                old_category = scalar_from_frontmatter(existing.frontmatter, "category")
            old_featured = scalar_from_frontmatter(existing.frontmatter, "featured")
            old_status = scalar_from_frontmatter(existing.frontmatter, "status")
            updates: dict = {
                "content_id": values["ID"],
                **classification,
                "featured": featured,
                "status": desired_status,
                "draft": draft_for_status(desired_status),
            }
            seo_before = existing.seo_complete
            seo_generated = False
            seo_failed_reason = ""
            generated: dict = {}
            seo_final_status = "已完成" if seo_before else seo_status
            if seo_status in {"待生成", "需重做"}:
                try:
                    generated = generate_seo_geo(
                        markdown_body_blocks(existing.body),
                        existing.title,
                        category_label,
                        existing.body,
                    )
                except Exception as exc:
                    seo_failed_reason = str(exc)
                    seo_final_status = "处理失败"
                    updates["seo_geo_status"] = "处理失败"
                    report.seo_processing_failed += 1
                    report.seo_failures.append(
                        {**row_ref(row_number, values), "文章类型": "旧文章", "原因": seo_failed_reason}
                    )
                else:
                    for key in SEO_FIELDS:
                        value = generated[key]
                        if seo_status == "需重做" or not seo_field_complete(existing.frontmatter, key):
                            updates[key] = value
                    updates["author"] = AUTHOR
                    updates["seo_geo_status"] = "已完成"
                    seo_generated = True
                    seo_final_status = "已完成"
            else:
                report.seo_preserved += 1
            changes = update_frontmatter_only(
                existing.path,
                updates,
                args.dry_run,
            )
            report.matched_old += 1
            if any(
                changes[key][0] != classification[key]
                for key in ("category", "subcategory", "content_type", "category_key", "category_label")
            ):
                report.old_category_updates += 1
            if changes["featured"][0] != featured:
                report.old_featured_updates += 1
            if changes["status"][0] != desired_status:
                report.old_status_updates += 1
            if seo_generated:
                report.seo_generated += 1
                report.seo_completed += 1
            elif seo_before:
                report.seo_completed += 1
            if not args.dry_run:
                set_excel(row_number, "是否精选", featured_text)
                set_excel(row_number, SEO_STATUS_COLUMN, seo_final_status)
            detail = {
                **row_ref(row_number, values),
                "结果": "dry-run旧文更新" if args.dry_run else "旧文已更新",
                "匹配方式": match.method,
                "原状态": values["状态"] or "（空）",
                "新状态": desired_status,
                "Excel原状态": values["状态"] or "空白",
                "标准化后状态": desired_status,
                "原分类": old_category or "（空）",
                "新分类": category_label,
                "原featured": old_featured if old_featured is not None else "（空）",
                "新featured": featured,
                "SEO/GEO字段是否完整": {"处理前": seo_before, "处理后": True if seo_generated else seo_before},
                "SEO/GEO最终内容": (
                    {key: generated[key] for key in SEO_FIELDS} if seo_generated else None
                ),
                "SEO/GEO质量校验": generated.get("_quality") if seo_generated else None,
                "SEO/GEO生成尝试": generated.get("_attempts") if seo_generated else [],
                "是否通过质量校验": generated["_quality"]["passed"] if seo_generated else None,
                "是否与正文第一句高度相似": generated["_quality"]["first_sentence_highly_similar"] if seo_generated else None,
                "SEO/GEO最终状态": seo_final_status,
                "SEO/GEO失败原因": seo_failed_reason,
                "是否会修改正文": "否",
                "是否会移动文件": "否",
                "Markdown": str(existing.path),
                "仅更新字段": {
                    key: {"原值": old, "新值": new}
                    for key, (old, new) in changes.items()
                    if not key.startswith("_")
                },
            }
            report.details.append(detail)
            report.generated_paths.append({**row_ref(row_number, values), "Markdown": str(existing.path)})
            continue

        if desired_status == "处理失败":
            report.skipped += 1
            report.details.append(
                {
                    **row_ref(row_number, values),
                    "结果": "跳过",
                    "原因": "Excel状态为处理失败且没有可保留的旧Markdown",
                    "新状态": desired_status,
                    "Excel原状态": values["状态"] or "空白",
                    "标准化后状态": desired_status,
                    "是否会修改正文": "否",
                    "是否会移动文件": "否",
                }
            )
            continue
        if not html_path:
            mark_failed(row_number, values, "HTML文件名无法精确匹配 .html 文件", "missing_html")
            continue
        final_article_status = "已发布" if desired_status == "待导入" else desired_status
        try:
            result = build_new_article(
                html_path,
                values["ID"],
                source_url,
                category_label,
                featured,
                final_article_status,
                site_root,
                args.dry_run,
                allow_remote_images=args.allow_remote_images,
            )
        except Exception as exc:
            mark_failed(row_number, values, str(exc))
            continue
        article_path = Path(result["article_path"]).resolve()
        if article_path in planned_article_paths:
            mark_failed(row_number, values, f"本次同步已虚拟创建相同Markdown：{article_path}")
            continue
        planned_article_paths.add(article_path)
        virtual_frontmatter = "\n".join(
            [
                f"content_id: {yaml_string(values['ID'])}",
                f"title: {yaml_string(result['title'])}",
                f"source_url: {yaml_string(result['source_url'])}",
            ]
        )
        register_article_in_indexes(
            ExistingArticle(
                article_path,
                values["ID"],
                result["title"],
                result["source_url"],
                virtual_frontmatter,
                "",
                result["seo_geo_final_status"] == "已完成",
                virtual=args.dry_run,
            ),
            by_id,
            by_url,
            by_title,
        )
        report.imported_new += 1
        if result["cover_used_placeholder"]:
            report.placeholder_covers += 1
        else:
            report.real_covers += 1
        if result["seo_geo_final_status"] == "已完成":
            report.seo_generated += 1
            report.seo_completed += 1
        else:
            report.seo_processing_failed += 1
            report.seo_failures.append(
                {
                    **row_ref(row_number, values),
                    "文章类型": "新文章",
                    "原因": result["seo_geo_error"],
                    "Markdown": str(result["article_path"]),
                }
            )
        if not args.dry_run:
            set_excel(row_number, "状态", final_article_status)
            set_excel(row_number, "是否精选", featured_text)
            set_excel(row_number, SEO_STATUS_COLUMN, result["seo_geo_final_status"])
        detail = {
            **row_ref(row_number, values),
            "结果": "dry-run新文导入" if args.dry_run else "新文已导入",
            "匹配方式": "未匹配旧文；按HTML首次导入",
            "原状态": values["状态"] or "（空）",
            "新状态": final_article_status,
            "Excel原状态": values["状态"] or "空白",
            "标准化后状态": desired_status,
            "导入成功后Excel状态": final_article_status,
            "Markdown写入状态": final_article_status,
            "Markdown写入draft": draft_for_status(final_article_status),
            "原分类": "（无旧Markdown）",
            "新分类": category_label,
            "原featured": "（无旧Markdown）",
            "新featured": featured,
            "SEO/GEO字段是否完整": {
                "处理前": False,
                "处理后": result["seo_geo_final_status"] == "已完成",
            },
            "SEO/GEO质量校验": result["quality_validation"],
            "SEO/GEO生成尝试": result["generation_attempts"],
            "是否通过质量校验": result["quality_validation"]["passed"],
            "是否与正文第一句高度相似": result["quality_validation"]["first_sentence_highly_similar"],
            "SEO/GEO最终状态": result["seo_geo_final_status"],
            "是否会修改正文": "否",
            "是否会移动文件": "否",
            **{key: str(value) if isinstance(value, Path) else value for key, value in result.items()},
        }
        report.details.append(detail)
        report.generated_paths.append(
            {
                **row_ref(row_number, values),
                "Markdown": str(result["article_path"]),
                "封面": str(result["cover_asset"]),
                "正文图片": result["image_paths"],
            }
        )

    if workbook_changed and not args.dry_run:
        save_workbook_atomic(workbook, excel_path)
    return report


def maintenance_records(args) -> list[dict]:
    site_root = args.site_root.resolve()
    excel_path = (site_root / args.excel).resolve()
    html_dir = (site_root / args.html_dir).resolve()
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = {clean_text(str(cell.value or "")): cell.column for cell in sheet[1]}
    rows_by_url: dict[str, dict] = {}
    for row_number in range(2, sheet.max_row + 1):
        values = {
            name: clean_text(str(sheet.cell(row_number, column).value or ""))
            for name, column in headers.items()
        }
        source_url = normalize_url(values.get("Binance链接"))
        if source_url:
            rows_by_url[source_url] = {"row": row_number, **values}
    requested = {
        token.strip().replace("\\", "/")
        for token in (args.paths or "").split(",")
        if token.strip()
    }
    records: list[dict] = []
    for path in sorted((site_root / "content" / "zh-cn").glob("**/*.md")):
        try:
            frontmatter, body, _newline = split_frontmatter(path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, ValueError):
            continue
        source_url = normalize_url(str(scalar_from_frontmatter(frontmatter, "source_url") or ""))
        if "binance.com/" not in source_url:
            continue
        relative = path.relative_to(site_root).as_posix()
        if requested and not any(
            token in {relative, path.name, path.stem} for token in requested
        ):
            continue
        row = rows_by_url.get(source_url, {})
        html_path = resolve_html(str(row.get("HTML文件名") or ""), html_dir)
        records.append(
            {
                "path": path,
                "relative_path": relative,
                "frontmatter": frontmatter,
                "body": body,
                "title": clean_text(str(scalar_from_frontmatter(frontmatter, "title") or "")),
                "category": clean_text(
                    str(
                        scalar_from_frontmatter(frontmatter, "category_label")
                        or scalar_from_frontmatter(frontmatter, "category")
                        or "Research"
                    )
                ),
                "source_url": source_url,
                "html_path": html_path,
                "excel_row": row.get("row"),
            }
        )
    if requested and len(records) != len(requested):
        found = {record["relative_path"] for record in records} | {
            record["path"].name for record in records
        } | {record["path"].stem for record in records}
        missing = sorted(requested - found)
        if missing:
            raise ValueError("--paths 中有文章无法匹配：" + ", ".join(missing))
    return records


def current_cover_info(record: dict, site_root: Path) -> dict:
    cover = str(scalar_from_frontmatter(record["frontmatter"], "cover") or "")
    asset_path = site_root / "assets" / cover.lstrip("/")
    static_path = site_root / "static" / cover.lstrip("/")
    exists = bool(cover) and (asset_path.is_file() or static_path.is_file())
    if not cover or not exists:
        cover_type = "缺失或路径失效"
    elif "placeholder" in cover.lower():
        cover_type = "占位封面"
    else:
        cover_type = "真实封面"
    return {
        "cover": cover,
        "cover_type": cover_type,
        "cover_exists": exists,
        "cover_status": clean_text(
            str(scalar_from_frontmatter(record["frontmatter"], "cover_status") or "")
        ),
    }


def inspect_recoverable_cover(record: dict, allow_remote: bool = True) -> dict:
    html_path = record["html_path"]
    if not html_path:
        return {"selection": None, "error": "找不到对应HTML"}
    try:
        document = html_path.read_text(encoding="utf-8")
        _article_html, article_start = find_article_html(document)
        selection = select_cover(
            document,
            article_start,
            record["title"],
            record["category"],
            html_path,
            allow_remote=allow_remote,
            collect_inventory=False,
        )
        return {"selection": selection, "error": ""}
    except Exception as exc:
        return {"selection": None, "error": str(exc)}


def run_cover_maintenance(args) -> dict:
    site_root = args.site_root.resolve()
    records = maintenance_records(args)
    details: list[dict] = []
    repaired = 0
    placeholders_regenerated = 0
    failures = 0
    for record in records:
        current = current_cover_info(record, site_root)
        needs_inspection = (
            current["cover_type"] != "真实封面" or args.recheck_real_covers
        )
        inspected = inspect_recoverable_cover(record, allow_remote=True) if needs_inspection else {
            "selection": None,
            "error": "",
        }
        selection: CoverSelection | None = inspected["selection"]
        if needs_inspection and selection is None and inspected["error"] == "找不到对应HTML":
            placeholder_data = generate_placeholder_cover(record["title"], record["category"])
            selection = CoverSelection(
                source="",
                source_label="统一新版占位封面",
                source_kind="placeholder",
                region="工作表中找不到对应HTML，无法恢复原始封面",
                candidate_count=0,
                resource_count=0,
                resource_types={},
                width=0,
                height=0,
                source_format="",
                is_true_top_visual=False,
                needs_normalization=False,
                output_width=1280,
                output_height=512,
                output_format="png",
                output_data=placeholder_data,
                used_remote=False,
                is_placeholder=True,
                cover_status="缺失",
                read_errors=[],
            )
        can_restore = bool(selection and not selection.is_placeholder)
        action = "保持现有真实封面"
        output_cover = current["cover"]
        source = "现有真实封面" if not needs_inspection else ""
        source_kind = "existing" if not needs_inspection else ""
        output_dimensions = ""
        body_hash_before = hashlib.sha256(record["body"].encode("utf-8")).hexdigest()

        if needs_inspection and selection:
            source = selection.source_label
            source_kind = selection.source_kind
            output_dimensions = f"{selection.output_width}x{selection.output_height}"
            preserve_existing_real = (
                current["cover_type"] == "真实封面" and selection.is_placeholder
            )
            if preserve_existing_real:
                output_cover = current["cover"]
                action = "未识别到更高置信度头图，保持现有真实封面"
            elif can_restore:
                name = f"{record['path'].stem}-cover.{selection.output_format}"
                output_cover = f"/images/covers/{name}"
                action = "恢复真实封面"
            else:
                name = f"{record['path'].stem}-cover-placeholder-v2.png"
                output_cover = f"/images/covers/{name}"
                action = "保留并重绘新版占位封面"

            if args.repair_covers and not args.dry_run and not preserve_existing_real:
                asset_path = site_root / "assets" / "images" / "covers" / name
                static_path = site_root / "static" / "images" / "covers" / name
                safe_write_binary(asset_path, selection.output_data)
                mirror_file(asset_path, static_path)
                update_frontmatter_only(
                    record["path"],
                    {
                        "cover": output_cover,
                        "cover_status": "正常" if can_restore else "缺失",
                    },
                    dry_run=False,
                )
            if args.repair_covers:
                if preserve_existing_real:
                    pass
                elif can_restore:
                    repaired += 1
                else:
                    placeholders_regenerated += 1
        elif needs_inspection:
            action = "无法检查，保持原封面"
            failures += 1

        details.append(
            {
                "文章标题": record["title"],
                "Markdown": record["relative_path"],
                "HTML文件名": record["html_path"].name if record["html_path"] else "",
                "当前封面": current["cover"],
                "当前封面类型": current["cover_type"],
                "封面来源": source,
                "封面来源类型": source_kind,
                "封面候选数量": selection.candidate_count if selection else 0,
                "封面所在HTML区域": selection.region if selection else "",
                "候选拒绝原因": selection.read_errors if selection else [],
                "是否能找到真实封面": (
                    can_restore if needs_inspection else current["cover_type"] == "真实封面"
                ),
                "修复后封面路径": output_cover,
                "原始尺寸": (
                    f"{selection.width}x{selection.height}" if selection and not selection.is_placeholder else ""
                ),
                "标准化后尺寸": output_dimensions,
                "是否使用占位封面": bool(
                    selection
                    and selection.is_placeholder
                    and current["cover_type"] != "真实封面"
                ),
                "动作": action,
                "错误": inspected["error"],
                "是否修改正文": "否",
                "是否修改URL": "否",
                "正文SHA256": body_hash_before,
            }
        )

    type_counts = {
        label: sum(current_cover_info(record, site_root)["cover_type"] == label for record in records)
        for label in ("真实封面", "占位封面", "缺失或路径失效")
    }
    return {
        "模式": "封面修复" if args.repair_covers else "封面审计",
        "dry_run": args.dry_run,
        "文章总数": len(records),
        "当前统计": type_counts,
        "可恢复真实封面": sum(item["是否能找到真实封面"] and item["当前封面类型"] != "真实封面" for item in details),
        "仍需占位封面": sum(item["是否使用占位封面"] for item in details),
        "实际恢复或计划恢复": repaired,
        "实际重绘或计划重绘占位封面": placeholders_regenerated,
        "检查失败": failures,
        "文章": details,
    }


def run_restore_preserved_covers(args) -> dict:
    """Restore preserved pre-recheck cover files without touching article content."""
    site_root = args.site_root.resolve()
    records = maintenance_records(args)
    details: list[dict] = []
    restored = 0
    kept = 0
    missing = 0
    for record in records:
        current = current_cover_info(record, site_root)
        action = "保持现有封面"
        restored_cover = current["cover"]
        if current["cover_type"] == "占位封面":
            preserved = None
            for extension in ("png", "jpg", "jpeg", "webp"):
                name = f"{record['path'].stem}-cover.{extension}"
                asset = site_root / "assets" / "images" / "covers" / name
                static = site_root / "static" / "images" / "covers" / name
                if asset.is_file() and static.is_file():
                    preserved = f"/images/covers/{name}"
                    break
            if preserved:
                restored_cover = preserved
                action = "恢复保留的原封面"
                restored += 1
                if not args.dry_run:
                    update_frontmatter_only(
                        record["path"],
                        {"cover": preserved, "cover_status": "正常"},
                        dry_run=False,
                    )
            else:
                action = "没有保留的原封面，继续使用占位封面"
                missing += 1
        else:
            kept += 1
        details.append(
            {
                "文章标题": record["title"],
                "Markdown": record["relative_path"],
                "当前封面": current["cover"],
                "恢复后封面": restored_cover,
                "动作": action,
                "是否修改正文": "否",
                "是否修改URL": "否",
            }
        )
    return {
        "模式": "恢复保留的原封面",
        "dry_run": args.dry_run,
        "文章总数": len(records),
        "恢复原封面": restored,
        "保持现有真实封面": kept,
        "仍需占位封面": missing,
        "文章": details,
    }


def run_adapt_covers(args) -> dict:
    """Adapt only contain-generated cover bitmaps; never touch Markdown."""
    site_root = args.site_root.resolve()
    requested = {
        token.strip().replace("\\", "/")
        for token in (args.paths or "").split(",")
        if token.strip()
    }
    records: list[dict] = []
    for path in sorted((site_root / "content").glob("**/*.md")):
        if path.name == "_index.md":
            continue
        try:
            frontmatter, body, _newline = split_frontmatter(path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, ValueError):
            continue
        cover = str(scalar_from_frontmatter(frontmatter, "cover") or "")
        if not cover:
            continue
        relative_path = path.relative_to(site_root).as_posix()
        if requested and not any(
            token in {relative_path, path.name, path.stem} for token in requested
        ):
            continue
        records.append({
            "path": path,
            "relative_path": relative_path,
            "frontmatter": frontmatter,
            "body": body,
            "title": clean_text(str(scalar_from_frontmatter(frontmatter, "title") or path.stem)),
        })
    details: list[dict] = []
    adapted_count = 0
    unchanged_count = 0
    skipped_count = 0
    backup_dir = site_root / "import" / "binance" / "cover-backup-before-fill"
    for record in records:
        current = current_cover_info(record, site_root)
        cover = current["cover"]
        if current["cover_type"] != "真实封面":
            skipped_count += 1
            details.append({
                "文章标题": record["title"],
                "封面": cover,
                "动作": "占位或失效封面不处理",
                "是否修改Markdown": "否",
            })
            continue
        relative = cover.removeprefix("/images/")
        asset_path = site_root / "assets" / "images" / relative
        static_path = site_root / "static" / "images" / relative
        if not asset_path.is_file():
            skipped_count += 1
            details.append({
                "文章标题": record["title"],
                "封面": cover,
                "动作": "assets封面不存在，跳过",
                "是否修改Markdown": "否",
            })
            continue
        current_data = asset_path.read_bytes()
        output_data, bbox = adapt_contain_cover_to_fill(current_data)
        backup_path = backup_dir / asset_path.name
        previously_adapted = backup_path.is_file() and "-cover-fill" not in asset_path.stem
        if output_data is None and not previously_adapted:
            unchanged_count += 1
            action = "原生5:2或无contain留白，保持不变"
        else:
            if output_data is None:
                output_data = current_data
            adapted_count += 1
            output_name = (
                asset_path.name
                if "-cover-fill" in asset_path.stem
                else f"{asset_path.stem}-fill.png"
            )
            output_cover = f"/images/covers/{output_name}"
            output_asset = site_root / "assets" / "images" / "covers" / output_name
            output_static = site_root / "static" / "images" / "covers" / output_name
            action = "铺满1280x512并发布新封面路径以刷新缓存"
            if not args.dry_run:
                backup_dir.mkdir(parents=True, exist_ok=True)
                if not backup_path.exists():
                    shutil.copy2(asset_path, backup_path)
                overwrite_binary_atomic(output_asset, output_data)
                overwrite_binary_atomic(output_static, output_data)
                update_frontmatter_only(
                    record["path"], {"cover": output_cover}, dry_run=False
                )
        details.append({
            "文章标题": record["title"],
            "封面": cover,
            "适配后封面": output_cover if output_data is not None else cover,
            "有效内容边界": list(bbox) if bbox else None,
            "输出尺寸": "1280x512",
            "动作": action,
            "是否修改Markdown": "仅cover字段" if output_data is not None else "否",
            "是否修改正文": "否",
            "是否修改URL": "否",
        })
    return {
        "模式": "封面视觉适配",
        "dry_run": args.dry_run,
        "文章总数": len(records),
        "需要适配": adapted_count,
        "保持不变": unchanged_count,
        "跳过": skipped_count,
        "备份目录": str(backup_dir),
        "文章": details,
    }


def current_summary_is_valid(summary: str, title: str, body: str) -> bool:
    first = markdown_first_sentence(body)
    blocks = markdown_body_blocks(body)
    pool = sentences(blocks)
    return bool(summary) and all(
        (
            40 <= visible_length(summary) <= 70,
            not contains_banned_language(summary),
            not has_abnormal_punctuation(summary),
            text_similarity(summary, first) < 0.72,
            text_similarity(summary, title) < 0.72,
            max((text_similarity(summary, item) for item in pool), default=0.0) < 0.82,
            body_supports_point(summary, " ".join(blocks)),
        )
    )


def run_summary_maintenance(args) -> dict:
    records = maintenance_records(args)
    details: list[dict] = []
    summary_repaired = 0
    summary_failed = 0
    points_repaired = 0
    points_failed = 0
    status_counts: dict[str, int] = {}
    for record in records:
        frontmatter = record["frontmatter"]
        body = record["body"]
        summary = clean_text(str(scalar_from_frontmatter(frontmatter, "summary") or ""))
        description = clean_text(str(scalar_from_frontmatter(frontmatter, "description") or ""))
        current_key_points = frontmatter_list_values(frontmatter, "key_points")
        seo_status = clean_text(str(scalar_from_frontmatter(frontmatter, "seo_geo_status") or "")) or "缺少状态"
        status_counts[seo_status] = status_counts.get(seo_status, 0) + 1
        first_sentence = markdown_first_sentence(body)
        summary_valid = current_summary_is_valid(summary, record["title"], body)
        pair_quality = (
            summary_copy_quality(summary, description, record["title"], body)
            if summary and description
            else None
        )
        description_valid = bool(description) and bool(
            pair_quality
            and not any(str(error).startswith("description") for error in pair_quality["errors"])
        )
        points_complete = 3 <= len(current_key_points) <= 5
        similarity = text_similarity(summary, first_sentence) if summary else 0.0
        legacy_source = "summary" if summary else ("description" if description else "正文.Hugo Summary")
        new_summary = summary
        new_description = description
        new_key_points = current_key_points
        summary_quality = pair_quality
        points_quality = core_points_quality(current_key_points, record["title"], body) if points_complete else None
        summary_error = ""
        points_error = ""
        actions: list[str] = []
        updates: dict = {}

        if args.repair_summaries and (not summary_valid or not description_valid):
            try:
                generated = generate_card_copy(
                    record["title"], record["category"], body
                )
                if not summary_valid:
                    new_summary = generated["summary"]
                    updates["summary"] = new_summary
                if not description_valid:
                    new_description = generated["description"]
                    updates["description"] = new_description
                summary_quality = summary_copy_quality(
                    new_summary, new_description, record["title"], body
                )
                actions.append("补齐独立summary/description")
                summary_repaired += 1
            except Exception as exc:
                summary_error = str(exc)
                if not summary_valid:
                    new_summary = ""
                    updates["summary"] = ""
                if not description_valid:
                    new_description = ""
                    updates["description"] = ""
                summary_quality = {
                    "passed": False,
                    "errors": [summary_error],
                    "first_sentence_highly_similar": False,
                }
                actions.append("summary/description生成失败，空值且不回退正文")
                summary_failed += 1
        elif not summary_valid or not description_valid:
            actions.append("summary/description需要修复")
        else:
            actions.append("保持现有有效summary/description")

        if args.repair_summaries and not points_complete:
            try:
                generated_points = generate_core_points(
                    record["title"], record["category"], body
                )
                new_key_points = generated_points["key_points"]
                points_quality = generated_points["quality"]
                updates["key_points"] = new_key_points
                actions.append("补齐3—5条核心观点")
                points_repaired += 1
            except Exception as exc:
                points_error = str(exc)
                new_key_points = []
                updates["key_points"] = []
                points_quality = {
                    "passed": False,
                    "errors": [points_error],
                    "count": 0,
                    "points": [],
                }
                actions.append("核心观点生成失败，保持空数组")
                points_failed += 1
        elif not points_complete:
            actions.append("核心观点需要修复")
        else:
            actions.append("保持现有核心观点")

        structured_complete = bool(
            new_summary
            and new_description
            and 3 <= len(new_key_points) <= 5
            and (summary_quality is None or summary_quality.get("passed", False))
            and (points_quality is None or points_quality.get("passed", False))
        )
        if args.repair_summaries:
            updates["summary_status"] = "已完成" if structured_complete else "处理失败"
            if not args.dry_run:
                update_frontmatter_only(record["path"], updates, dry_run=False)

        details.append(
            {
                "文章标题": record["title"],
                "Markdown": record["relative_path"],
                "当前summary": summary,
                "正文第一句": first_sentence,
                "summary与正文第一句相似度": round(similarity, 4),
                "当前summary是否有效": summary_valid,
                "当前description是否有效": description_valid,
                "当前核心观点": current_key_points,
                "当前核心观点数量": len(current_key_points),
                "新summary": new_summary,
                "新description": new_description,
                "新核心观点": new_key_points,
                "summary/description质量校验": summary_quality,
                "核心观点质量校验": points_quality,
                "seo_geo_status": seo_status,
                "summary_status": "已完成" if structured_complete else "处理失败",
                "修复前前台来源": legacy_source,
                "修复后前台将读取": "summary" if new_summary else ("description" if new_description else "空字符串"),
                "动作": "；".join(actions),
                "摘要错误": summary_error,
                "核心观点错误": points_error,
                "是否修改正文": "否",
                "是否修改URL": "否",
                "正文SHA256": hashlib.sha256(body.encode("utf-8")).hexdigest(),
            }
        )

    return {
        "模式": "摘要修复" if args.repair_summaries else "摘要审计",
        "dry_run": args.dry_run,
        "文章总数": len(records),
        "有效summary": sum(item["当前summary是否有效"] for item in details),
        "summary为空": sum(not item["当前summary"] for item in details),
        "核心观点完整": sum(3 <= item["当前核心观点数量"] <= 5 for item in details),
        "核心观点为空": sum(item["当前核心观点数量"] == 0 for item in details),
        "summary与正文第一句高度相似": sum(item["summary与正文第一句相似度"] >= 0.72 for item in details),
        "修复前前台回退正文": sum(item["修复前前台来源"] == "正文.Hugo Summary" for item in details),
        "SEO/GEO状态": status_counts,
        "summary/description修复成功": summary_repaired,
        "summary/description修复失败": summary_failed,
        "核心观点修复成功": points_repaired,
        "核心观点修复失败": points_failed,
        "修复后结构完整": sum(item["summary_status"] == "已完成" for item in details),
        "文章": details,
    }


def run_key_points_maintenance(args) -> dict:
    records = maintenance_records(args)
    details: list[dict] = []
    repaired = 0
    failed = 0
    for record in records:
        current = frontmatter_list_values(record["frontmatter"], "key_points")
        complete = 3 <= len(current) <= 5
        generated = current
        quality = core_points_quality(current, record["title"], record["body"]) if complete else None
        error = ""
        action = "保持现有核心观点" if complete else "需要补齐核心观点"
        if args.repair_key_points and not complete:
            try:
                result = generate_core_points(record["title"], record["category"], record["body"])
                generated = result["key_points"]
                quality = result["quality"]
                action = "补齐3—5条核心观点"
                repaired += 1
                if not args.dry_run:
                    update_frontmatter_only(
                        record["path"],
                        {
                            "key_points": generated,
                            "key_points_status": "已完成",
                        },
                        dry_run=False,
                    )
            except Exception as exc:
                error = str(exc)
                generated = []
                quality = {"passed": False, "errors": [error], "count": 0, "points": []}
                action = "核心观点生成失败，保持空数组"
                failed += 1
                if not args.dry_run:
                    update_frontmatter_only(
                        record["path"],
                        {
                            "key_points": [],
                            "key_points_status": "处理失败",
                        },
                        dry_run=False,
                    )
        details.append(
            {
                "文章标题": record["title"],
                "Markdown": record["relative_path"],
                "当前核心观点数量": len(current),
                "当前核心观点": current,
                "新核心观点数量": len(generated),
                "新核心观点": generated,
                "质量校验": quality,
                "动作": action,
                "错误": error,
                "是否修改summary": "否",
                "是否修改description": "否",
                "是否修改正文": "否",
                "是否修改URL": "否",
                "正文SHA256": hashlib.sha256(record["body"].encode("utf-8")).hexdigest(),
            }
        )
    return {
        "模式": "核心观点修复" if args.repair_key_points else "核心观点审计",
        "dry_run": args.dry_run,
        "文章总数": len(records),
        "原有核心观点完整": sum(3 <= item["当前核心观点数量"] <= 5 for item in details),
        "原有核心观点为空": sum(item["当前核心观点数量"] == 0 for item in details),
        "修复成功": repaired,
        "修复失败": failed,
        "修复后核心观点完整": sum(3 <= item["新核心观点数量"] <= 5 for item in details),
        "文章": details,
    }


def run_missing_summary_maintenance(args) -> dict:
    records = maintenance_records(args)
    details: list[dict] = []
    repaired = 0
    failed = 0
    for record in records:
        summary = clean_text(str(scalar_from_frontmatter(record["frontmatter"], "summary") or ""))
        description = clean_text(str(scalar_from_frontmatter(record["frontmatter"], "description") or ""))
        points = frontmatter_list_values(record["frontmatter"], "key_points")
        new_summary = summary
        new_description = description
        quality = None
        error = ""
        action = "保持现有摘要" if summary else "需要补齐摘要"
        if args.repair_missing_summaries and not summary:
            try:
                result = generate_missing_summary(record["path"], record["title"], record["body"], points)
                new_summary = result["summary"]
                new_description = result["description"]
                quality = result["quality"]
                action = "补齐经审定的summary和description"
                repaired += 1
                if not args.dry_run:
                    update_frontmatter_only(
                        record["path"],
                        {
                            "summary": new_summary,
                            "description": new_description,
                            "summary_status": "已完成",
                        },
                        dry_run=False,
                    )
            except Exception as exc:
                error = str(exc)
                action = "生成失败，保持空值且前台不回退正文"
                failed += 1
                if not args.dry_run:
                    update_frontmatter_only(
                        record["path"],
                        {"summary_status": "处理失败"},
                        dry_run=False,
                    )
        details.append(
            {
                "文章标题": record["title"],
                "Markdown": record["relative_path"],
                "当前summary": summary,
                "新summary": new_summary,
                "新description": new_description,
                "核心观点数量": len(points),
                "质量校验": quality,
                "动作": action,
                "错误": error,
                "是否修改核心观点": "否",
                "是否修改正文": "否",
                "是否修改URL": "否",
                "正文SHA256": hashlib.sha256(record["body"].encode("utf-8")).hexdigest(),
            }
        )
    return {
        "模式": "缺失摘要修复" if args.repair_missing_summaries else "缺失摘要审计",
        "dry_run": args.dry_run,
        "文章总数": len(records),
        "原有summary": sum(bool(item["当前summary"]) for item in details),
        "原有summary为空": sum(not item["当前summary"] for item in details),
        "修复成功": repaired,
        "修复失败": failed,
        "修复后summary完整": sum(bool(item["新summary"]) for item in details),
        "文章": details,
    }


def parse_args():
    parser = argparse.ArgumentParser(description="从 Excel 总控表同步 Binance 历史文章")
    parser.add_argument("--site-root", type=Path, default=Path.cwd())
    parser.add_argument("--excel", type=Path, default=Path("import/binance/content.xlsx"))
    parser.add_argument("--html-dir", type=Path, default=Path("import/binance/html"))
    selection = parser.add_mutually_exclusive_group(required=True)
    selection.add_argument("--rows", help="仅处理指定 Excel 行号，例如 2,3,53")
    selection.add_argument("--all", action="store_true", help="显式处理整张表")
    selection.add_argument("--normalize-excel", action="store_true", help="仅标准化Excel控制字段，不同步Markdown")
    selection.add_argument("--audit-covers", action="store_true", help="审计已导入文章封面，不修改正文")
    selection.add_argument("--repair-covers", action="store_true", help="独立回填真实封面，不重新导入正文")
    selection.add_argument(
        "--restore-preserved-covers",
        action="store_true",
        help="恢复被复检流程替换前仍保存在assets/static中的原封面",
    )
    selection.add_argument(
        "--adapt-covers",
        action="store_true",
        help="仅将contain留白封面按cover方式适配到1280x512，不修改Markdown",
    )
    parser.add_argument(
        "--recheck-real-covers",
        action="store_true",
        help="复检当前标记为真实封面的文章，识别误用的正文图片",
    )
    selection.add_argument("--audit-summaries", action="store_true", help="审计已导入文章摘要和前台回退来源")
    selection.add_argument("--repair-summaries", action="store_true", help="独立修复摘要、描述与核心观点，不重新导入正文")
    selection.add_argument("--audit-key-points", action="store_true", help="审计已导入文章核心观点，不修改正文")
    selection.add_argument("--repair-key-points", action="store_true", help="仅补齐全部缺失核心观点，不修改摘要或正文")
    selection.add_argument("--audit-missing-summaries", action="store_true", help="审计summary为空的文章")
    selection.add_argument("--repair-missing-summaries", action="store_true", help="仅补齐全部空summary和description，不覆盖现有摘要")
    parser.add_argument(
        "--paths",
        help="维护模式仅处理指定Markdown路径、文件名或stem，多个值用逗号分隔",
    )
    parser.add_argument("--dry-run", action="store_true", help="只输出计划，不写任何文件")
    parser.add_argument(
        "--allow-remote-images",
        action="store_true",
        help="允许在离线资源缺失时最后尝试远程图片；默认完全离线",
    )
    return parser.parse_args()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    args = parse_args()
    try:
        if args.adapt_covers:
            report = run_adapt_covers(args)
        elif args.restore_preserved_covers:
            report = run_restore_preserved_covers(args)
        elif args.audit_covers or args.repair_covers:
            report = run_cover_maintenance(args)
        elif args.audit_missing_summaries or args.repair_missing_summaries:
            report = run_missing_summary_maintenance(args)
        elif args.audit_key_points or args.repair_key_points:
            report = run_key_points_maintenance(args)
        elif args.audit_summaries or args.repair_summaries:
            report = run_summary_maintenance(args)
        else:
            report = run_sync(args)
    except PreflightValidationError as exc:
        print(
            json.dumps(
                {
                    "同步前检查": "失败，未处理任何文章",
                    "新增文章数量": 0,
                    "更新文章数量": 0,
                    "跳过数量": 0,
                    "错误数量": len(exc.errors),
                    "错误明细": exc.errors,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 2
    except Exception as exc:
        print(json.dumps({"fatal": str(exc)}, ensure_ascii=False, indent=2))
        return 1
    output_data = report.as_dict() if isinstance(report, SyncReport) else report
    output = json.dumps(output_data, ensure_ascii=False, indent=2)
    print(output)
    is_audit = (
        args.audit_covers
        or args.audit_summaries
        or args.audit_key_points
        or args.audit_missing_summaries
    )
    if not args.dry_run and not is_audit:
        log_path = args.site_root.resolve() / "logs" / "content-sync.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now(ZoneInfo("Asia/Shanghai")).isoformat(timespec="seconds")
        with log_path.open("a", encoding="utf-8", newline="\n") as handle:
            handle.write(f"[{timestamp}]\n{output}\n")
    if isinstance(report, SyncReport):
        return 0 if not report.failures else 2
    return 0 if not report.get("检查失败", 0) else 2


if __name__ == "__main__":
    sys.exit(main())
